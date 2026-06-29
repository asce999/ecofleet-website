import os
import django
import openpyxl
from decimal import Decimal

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ecofleet.settings')
django.setup()

from core.attendance import calculate_salary_data
from core.models import SalaryConfig, EmployeeSalaryOverride

btpl_file = r"C:\Users\ardcr\Documents\BTPL_CV_DISTRIBUTIONS_PAYMENT_FEB-2026.xlsx"
wb = openpyxl.load_workbook(btpl_file, data_only=True)
ws = wb.active

test_cases = [
    "AMOL VAIBHAV MASKAR",
    "PREM DATTU RANDIVE",
    "PRATIK PRAMOD KADAM",
    "MAYUR ABHIMAN THORAT",
    "SAHIL MANOJ JADHAV",
    "RAVI SALIM GAIKWAD"
]

config = SalaryConfig.get_solo()

employees_data = []

# Header is on row 4
for r in range(5, ws.max_row + 1):
    name_val = ws.cell(row=r, column=8).value
    if not name_val: continue
    name = str(name_val).strip()
    
    if name in test_cases:
        dept = str(ws.cell(row=r, column=3).value or "")
        emp_id = str(ws.cell(row=r, column=4).value or "")
        esic_no = str(ws.cell(row=r, column=5).value or "")
        pf_no = str(ws.cell(row=r, column=6).value or "")
        uan_no = str(ws.cell(row=r, column=7).value or "")
        
        payable_days = Decimal(str(ws.cell(row=r, column=14).value or 0))
        extra_days = Decimal(str(ws.cell(row=r, column=15).value or 0))
        
        adhoc_sal_inc = Decimal(str(ws.cell(row=r, column=17).value or 0))
        basic = Decimal(str(ws.cell(row=r, column=16).value or 0))
        if basic > 0:
            adhoc_inc_pct = (adhoc_sal_inc / basic) * Decimal('100.0')
        else:
            adhoc_inc_pct = Decimal('0')
            
        adhoc_allow_amount = Decimal(str(ws.cell(row=r, column=19).value or 0))
        if payable_days > 0 and adhoc_allow_amount > 0:
            adhoc_allow_monthly = (adhoc_allow_amount * Decimal('26') / payable_days).quantize(Decimal('1'))
        else:
            adhoc_allow_monthly = Decimal('0')
            
        advance = Decimal(str(ws.cell(row=r, column=29).value or 0))
        lwf = Decimal(str(ws.cell(row=r, column=30).value or 0))
        other_deduction = Decimal(str(ws.cell(row=r, column=32).value or 0))
        
        # Save overrides so the engine uses them
        EmployeeSalaryOverride.objects.update_or_create(
            employee_name=name,
            defaults={
                'adhoc_salary_increase_pct': adhoc_inc_pct,
                'adhoc_allowance_monthly_amount': adhoc_allow_monthly,
                'advance': advance,
                'lwf': lwf,
                'other_deduction': other_deduction
            }
        )
        
        # We need to test the engine's output vs sheet's output
        sheet_total_a = Decimal(str(ws.cell(row=r, column=20).value or 0))
        sheet_total_b = Decimal(str(ws.cell(row=r, column=25).value or 0))
        sheet_pf = Decimal(str(ws.cell(row=r, column=26).value or 0))
        sheet_esic = Decimal(str(ws.cell(row=r, column=27).value or 0))
        sheet_pt = Decimal(str(ws.cell(row=r, column=28).value or 0))
        sheet_net = Decimal(str(ws.cell(row=r, column=34).value or 0))
        
        employees_data.append({
            'name': name,
            'department': dept,
            'emp_id': emp_id,
            'esic_no': esic_no,
            'pf_no': pf_no,
            'uan_no': uan_no,
            'payable_days': float(payable_days),
            'extra_days': float(extra_days),
            'attendance': {},
            
            # Expected values for comparison
            '_expected_total_a': sheet_total_a,
            '_expected_total_b': sheet_total_b,
            '_expected_pf': sheet_pf,
            '_expected_esic': sheet_esic,
            '_expected_pt': sheet_pt,
            '_expected_net': sheet_net
        })

attendance_data = {
    'year': 2026,
    'month': 2,
    'employees': employees_data
}

result = calculate_salary_data(attendance_data)

print(f"{'Employee Name':<25} | {'Metric':<10} | {'Expected':<10} | {'Calculated':<10} | {'Diff':<10}")
print("-" * 75)

for i, emp_res in enumerate(result['employees']):
    orig = employees_data[i]
    
    metrics = [
        ('Total(A)', orig['_expected_total_a'], emp_res['total_a']),
        ('Total(B)', orig['_expected_total_b'], emp_res['total_b']),
        ('PF', orig['_expected_pf'], emp_res['pf']),
        ('ESIC', orig['_expected_esic'], emp_res['esic_employee']),
        ('PT Tax', orig['_expected_pt'], emp_res['pt']),
        ('Net Pymt', orig['_expected_net'], emp_res['net_payment']),
    ]
    
    for metric, exp, calc in metrics:
        diff = calc - exp
        diff_str = str(diff) if diff != 0 else ""
        print(f"{orig['name']:<25} | {metric:<10} | {exp:<10.2f} | {calc:<10.2f} | {diff_str:<10}")
    print("-" * 75)
