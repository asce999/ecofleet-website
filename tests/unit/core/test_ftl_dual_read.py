from django.test import TestCase, Client
from django.urls import reverse
from core.models import UserProfile, MigrationFeatureFlags, Shipment, FtlWorkbook
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
import tempfile
import os
import openpyxl
from datetime import date

class FtlDualReadTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testadmin', password='password123', is_staff=True)
        profile = UserProfile.objects.get(user=self.user)
        profile.role = 'Director'
        profile.can_use_ftl = True
        profile.save()
        self.client = Client()
        self.client.force_login(self.user)
        
        self.temp_dir = tempfile.TemporaryDirectory()
        self.excel_path = os.path.join(self.temp_dir.name, 'test_active.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value='Date of Booking')
        ws.cell(row=1, column=2, value='LR No.')
        ws.cell(row=2, column=1, value='2026-06-01')
        ws.cell(row=2, column=2, value='LR999')
        wb.save(self.excel_path)
        wb.close()
        
        with open(self.excel_path, 'rb') as f:
            self.wb = FtlWorkbook.objects.create(
                original_name='test.xlsx',
                active_sheet='Sheet1',
                is_active=True,
                file=SimpleUploadedFile('test.xlsx', f.read())
            )
            
        self.shipment = Shipment.objects.create(
            shipment_type='FTL',
            source_key='LR999|2026-06-01',
            dispatch_date=date(2026, 6, 1),
            origin='Delhi',
            metadata={'lr_number': 'LR999'}
        )

    def tearDown(self):
        self.temp_dir.cleanup()
        if self.wb.file and os.path.exists(self.wb.file.path):
            os.remove(self.wb.file.path)

    def test_dual_read_off_uses_excel(self):
        flags = MigrationFeatureFlags.get_solo()
        flags.use_database_reads = False
        flags.save()
        
        resp = self.client.get(reverse('ftl_sheet'))
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn('Delhi', str(resp.content))

    def test_dual_read_on_uses_db(self):
        flags = MigrationFeatureFlags.get_solo()
        flags.use_database_reads = True
        flags.save()
        
        resp = self.client.get(reverse('ftl_sheet'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('Delhi', str(resp.content))
        self.assertIn('preview', resp.context)
        self.assertEqual(resp.context['preview']['total_rows'], 1)
