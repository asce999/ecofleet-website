import openpyxl
import os
import tempfile
from django.test import TestCase
from core.models import ImportJob, Shipment, ShipmentStatus
from core.importers.excel_importer import ExcelImporter

class ImportersTest(TestCase):
    def setUp(self):
        # Create a temporary excel file
        self.fd, self.temp_path = tempfile.mkstemp(suffix='.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Headers 
        headers = ["Date of Booking", "Vehicle Number", "LR Number", "From Location", "To Location"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
            
        # Row 2
        ws.cell(row=2, column=1, value="2026-06-15") # Date of Booking
        ws.cell(row=2, column=2, value="MH12AB1234") # Vehicle
        ws.cell(row=2, column=3, value="LR001")      # LR
        ws.cell(row=2, column=4, value="Pune")       # From
        ws.cell(row=2, column=5, value="Mumbai")     # To
        
        # Row 3 (Duplicate row to test same file imported twice)
        ws.cell(row=3, column=1, value="2026-06-16") 
        ws.cell(row=3, column=2, value="MH12AB5678") 
        ws.cell(row=3, column=3, value="LR002")      
        ws.cell(row=3, column=4, value="Pune")       
        ws.cell(row=3, column=5, value="Delhi")      
        
        wb.save(self.temp_path)
        wb.close()
        
    def tearDown(self):
        os.close(self.fd)
        os.remove(self.temp_path)

    def test_ftl_importer_idempotency(self):
        importer = ExcelImporter()
        job1 = ImportJob.objects.create(workbook_type='FTL')
        
        # First import
        importer.process_ftl_workbook(job1.id, self.temp_path)
        
        job1.refresh_from_db()
        self.assertEqual(job1.status, 'COMPLETED')
        self.assertEqual(job1.failed_rows, 0)
        
        shipments_count = Shipment.objects.count()
        self.assertEqual(shipments_count, 2)
        # One status row per shipment after the first import.
        self.assertEqual(ShipmentStatus.objects.count(), 2)

        # Second import (double import)
        job2 = ImportJob.objects.create(workbook_type='FTL')
        importer.process_ftl_workbook(job2.id, self.temp_path)

        job2.refresh_from_db()
        self.assertEqual(job2.status, 'COMPLETED')
        self.assertEqual(job2.failed_rows, 0)

        # Count should still be 2, not 4
        self.assertEqual(Shipment.objects.count(), 2)
        # Status log must not bloat on re-import (F-02): still 2, not 4.
        self.assertEqual(ShipmentStatus.objects.count(), 2)
