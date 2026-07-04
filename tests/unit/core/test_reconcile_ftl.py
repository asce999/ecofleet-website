import os
import tempfile
import openpyxl
from io import StringIO
from datetime import date
from django.test import TestCase
from django.core.management import call_command
from core.models import Shipment, Vehicle


class ReconcileFtlTest(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.excel_path = os.path.join(self.temp_dir.name, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        headers = ['Date of Booking', 'LR No.', 'Vehicle No', 'From', 'To']
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
            
        ws.cell(row=2, column=1, value='2026-06-01')
        ws.cell(row=2, column=2, value='LR100')
        ws.cell(row=2, column=3, value='TN01AB1234')
        ws.cell(row=2, column=4, value='Chennai')
        ws.cell(row=2, column=5, value='Bangalore')
        
        wb.save(self.excel_path)
        wb.close()
        
        self.vehicle = Vehicle.objects.create(registration_number='TN01AB1234')
        self.shipment = Shipment.objects.create(
            shipment_type='FTL',
            source_key='LR100|2026-06-01',
            dispatch_date=date(2026, 6, 1),
            origin='Chennai',
            destination='Bangalore',
            vehicle=self.vehicle,
            metadata={'lr_number': 'LR100'}
        )

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_reconcile_success(self):
        out = StringIO()
        try:
            call_command("reconcile_ftl", workbook_path=self.excel_path, stdout=out)
        except SystemExit as e:
            self.assertEqual(e.code, 0)
        self.assertIn("Reconciliation SUCCESSFUL", out.getvalue())

    def test_reconcile_failure_mismatched_field(self):
        self.shipment.origin = 'Mumbai'
        self.shipment.save()
        
        out = StringIO()
        try:
            call_command("reconcile_ftl", workbook_path=self.excel_path, stdout=out)
        except SystemExit as e:
            self.assertEqual(e.code, 1)
        self.assertIn("Reconciliation FAILED", out.getvalue())
        self.assertIn("origin: 'Chennai' != 'Mumbai'", out.getvalue())
