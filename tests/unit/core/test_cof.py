import os
import time
import tempfile
import threading
from unittest import mock
from django.test import TestCase, override_settings
from openpyxl import Workbook

from core import cof
from django.conf import settings

class COFTestCase(TestCase):
    def setUp(self):
        # Create a temporary directory for locking and templates
        self.test_dir = tempfile.TemporaryDirectory()
        self.lock_path = os.path.join(self.test_dir.name, 'cof.lock')
        self.letterhead_path = os.path.join(self.test_dir.name, 'letterhead.docx')
        
        # Create a dummy letterhead docx file so generate_cof doesn't fail on AssetMissing
        from docx import Document
        doc = Document()
        doc.add_paragraph("DUMMY LETTERHEAD")
        doc.save(self.letterhead_path)

        # Create a dummy COF workbook
        self.wb_path = os.path.join(self.test_dir.name, 'cof_test.xlsx')
        self._build_dummy_workbook(self.wb_path)

    def tearDown(self):
        self.test_dir.cleanup()

    def _build_dummy_workbook(self, path, serial=5, efe_number=10):
        wb = Workbook()
        # Data Sheet
        ws = wb.active
        ws.title = cof.DATA_SHEET
        # headers
        ws.append(["#", "LR Number", "Pickup Date", "Invoice Number", "Remark",
                   "Dealer", "State", "Claim Amount", "CN Amount", "Delhivery Doc",
                   "Doc Date", "Status Delhivery", "Remarks", "COF Date",
                   "Status Optlog", "Ref Delhivery"])
        
        # previous row to simulate next serial/efe
        ws.append([serial, "LR123", "2026-01-01", "INV123", "Short", "DealerA", "StateA",
                   100, "", "", "", "Pending", "", "2026-01-02", f"EFE-{efe_number:03d}", ""])
        
        # Template sheet
        # Assuming settings.COF_TEMPLATE_SHEET defaults to something, we can override or just use 'Template'
        template_sheet = getattr(settings, 'COF_TEMPLATE_SHEET', 'Template')
        ws_template = wb.create_sheet(template_sheet)
        ws_template["A1"] = "COF TEMPLATE DUMMY"
        
        wb.save(path)
        wb.close()

    def test_get_next_cof_info(self):
        """1. get_next_cof_info: Verify correct next COF serial and EFE returned"""
        info = cof.get_next_cof_info(self.wb_path)
        self.assertEqual(info["serial"], 6)
        self.assertEqual(info["claim_number"], 11)
        self.assertEqual(info["cof_number"], "EFE-COF-0011")
        self.assertEqual(info["optlog"], "EFE-011")

    @override_settings()
    def test_workbook_locking(self):
        """2. Workbook Locking: First lock succeeds, second blocks."""
        settings.COF_LOCK_PATH = self.lock_path
        
        # First lock succeeds
        with cof.workbook_lock(timeout=1):
            self.assertTrue(os.path.exists(self.lock_path))
            
            # Second acquisition attempt blocks/fails
            with self.assertRaises(cof.COFLockTimeout):
                with cof.workbook_lock(timeout=0.1, poll=0.05):
                    pass

    @override_settings()
    def test_stale_lock_cleanup(self):
        """3. Stale Lock Cleanup: Stale lock removed, acquisition succeeds."""
        settings.COF_LOCK_PATH = self.lock_path
        
        # Create a lock file and make it stale by modifying its mtime
        os.makedirs(os.path.dirname(self.lock_path), exist_ok=True)
        with open(self.lock_path, 'w') as f:
            f.write("stale lock")
        
        # Set mtime to 10 minutes ago
        stale_time = time.time() - 600
        os.utime(self.lock_path, (stale_time, stale_time))

        # This should succeed and remove the stale lock
        with cof.workbook_lock(timeout=1, stale=120):
            self.assertTrue(os.path.exists(self.lock_path))
            # The newly acquired lock should be fresh
            self.assertGreater(os.path.getmtime(self.lock_path), stale_time + 500)

    @mock.patch('core.cof.load_workbook')
    def test_workbook_in_use(self, mock_load):
        """4. WorkbookInUse: Simulate locked workbook"""
        # Make load_workbook raise PermissionError
        mock_load.side_effect = PermissionError("Permission denied")

        d = {
            "consignee_name": "Test Consignee",
            "consignee_address": "Test Addr",
            "consignee_state": "Test State",
            "invoice_numbers": "INV-001",
            "invoice_date": "2026-06-01",
            "num_packages": 2,
            "weight": 10.5,
            "destination_city": "Test City",
            "delivery_date": "2026-06-10",
            "lr_number": "LR-001",
            "loss_amount": 500,
            "remark": "Short",
            "pickup_date": "2026-05-30",
            "dealer_name": "Test Dealer",
            "state": "Test State",
            "status_delhivery": "Delivered",
            "ref_delhivery": "Ref-123"
        }

        with self.assertRaises(cof.WorkbookInUse):
            cof.generate_cof(d, self.wb_path)

    @override_settings()
    def test_generate_cof_end_to_end(self):
        """5. generate_cof End-to-End"""
        settings.COF_LOCK_PATH = self.lock_path
        settings.COF_LETTERHEAD_PATH = self.letterhead_path
        settings.COF_TEMPLATE_SHEET = getattr(settings, 'COF_TEMPLATE_SHEET', 'Template')

        d = {
            "consignee_name": "Test Consignee",
            "consignee_address": "Test Addr",
            "consignee_state": "Test State",
            "invoice_numbers": "INV-001",
            "invoice_date": "2026-06-01",
            "num_packages": 2,
            "weight": 10.5,
            "destination_city": "Test City",
            "delivery_date": "2026-06-10",
            "lr_number": "LR-001",
            "loss_amount": 500,
            "remark": "Short",
            "pickup_date": "2026-05-30",
            "dealer_name": "Test Dealer",
            "state": "Test State",
            "status_delhivery": "Delivered",
            "ref_delhivery": "Ref-123"
        }

        res = cof.generate_cof(d, self.wb_path)

        # Verify return payload
        self.assertEqual(res["serial"], 6)
        self.assertEqual(res["claim_number"], 11)
        self.assertEqual(res["cof_number"], "EFE-COF-0011")
        self.assertTrue(res["docx"].getbuffer().nbytes > 0)

        # Re-verify workbook state
        from openpyxl import load_workbook
        wb = load_workbook(self.wb_path)
        ws = wb[cof.DATA_SHEET]

        # The new row should be row 3
        row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        self.assertEqual(row3[0], 6) # serial
        self.assertEqual(row3[14], "EFE-011") # optlog

        # Verify a new sheet was created for the COF
        self.assertIn("Test Consignee", wb.sheetnames)
        wb.close()
