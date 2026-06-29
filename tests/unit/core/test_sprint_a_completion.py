from django.test import TestCase
from django.contrib.auth.models import User
from core.models import UserProfile
from core.utils.excel import sanitize_excel_formula
from core.operations.services.insights import InsightsService
from core.services.sheet_parser import get_sheet_names
import os
import tempfile
import openpyxl

class SprintACompletionTests(TestCase):
    def test_sanitize_excel_formula(self):
        self.assertEqual(sanitize_excel_formula('=cmd()'), "'=cmd()")
        self.assertEqual(sanitize_excel_formula('+cmd()'), "'+cmd()")
        self.assertEqual(sanitize_excel_formula('-cmd()'), "'-cmd()")
        self.assertEqual(sanitize_excel_formula('@cmd()'), "'@cmd()")
        self.assertEqual(sanitize_excel_formula('Normal text'), 'Normal text')
        self.assertEqual(sanitize_excel_formula(None), None)

    def test_least_privilege_defaults(self):
        user = User.objects.create_user(username='test_user')
        profile = user.profile
        self.assertFalse(profile.can_use_attendance)
        self.assertFalse(profile.can_use_btpl)
        self.assertFalse(profile.can_use_cof)
        self.assertFalse(profile.can_use_ftl)
        self.assertFalse(profile.can_use_morning)
        self.assertFalse(profile.can_use_pendency)
        self.assertFalse(profile.can_use_prev_month)

    def test_presentation_separation(self):
        score = InsightsService.calculate_operational_score({}, False, False)
        self.assertIsNone(score.availability)

    def test_workbook_sheet_discovery(self):
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet2')
        wb.create_sheet('Sheet3')
        
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        wb.save(tmp_path)
        try:
            sheets = get_sheet_names(tmp_path)
            self.assertIn('Sheet', sheets)
            self.assertIn('Sheet2', sheets)
            self.assertIn('Sheet3', sheets)
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
