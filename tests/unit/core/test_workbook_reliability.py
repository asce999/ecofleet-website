import os
import time
from django.test import TestCase
from django.conf import settings
from openpyxl import Workbook
from core.workbook.locking import workbook_lock, LockTimeoutError
from core.workbook.helpers import atomic_save_workbook
import tempfile

class WorkbookReliabilityTests(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.tmp_path = self.temp_dir.name
        
    def tearDown(self):
        self.temp_dir.cleanup()

    def test_workbook_lock_success(self):
        wb_path = os.path.join(self.tmp_path, "test.xlsx")
        with workbook_lock(wb_path):
            self.assertTrue(os.path.exists(f"{wb_path}.lock"))
        self.assertFalse(os.path.exists(f"{wb_path}.lock"))

    def test_workbook_lock_timeout_stale(self):
        wb_path = os.path.join(self.tmp_path, "test.xlsx")
        lock_path = f"{wb_path}.lock"
        
        # Create a stale lock file
        with open(lock_path, 'w') as f:
            f.write("999999") # An unlikely PID
            
        # Simulate TTL expiration by setting mtime in the past
        os.utime(lock_path, (time.time() - 200, time.time() - 200))
        
        # This should succeed and remove the stale lock
        with workbook_lock(wb_path):
            self.assertTrue(os.path.exists(lock_path))
        self.assertFalse(os.path.exists(lock_path))

    def test_atomic_save_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test Data"
        
        target_path = os.path.join(self.tmp_path, "atomic_test.xlsx")
        atomic_save_workbook(wb, target_path)
        
        self.assertTrue(os.path.exists(target_path))
        
        # Load and verify
        from openpyxl import load_workbook
        loaded_wb = load_workbook(target_path)
        loaded_ws = loaded_wb.active
        self.assertEqual(loaded_ws['A1'].value, "Test Data")
