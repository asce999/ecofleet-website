"""
Pendency (delayed shipments) report — upload model.
The employee uploads the 2W and CV master reports; we extract In-Transit
shipments (optionally only those delayed by >= a chosen number of days, or
everything in transit), and build a workbook with a 2W sheet and a CV sheet.
The report month is read from the uploaded file names.
"""
import io
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportError(Exception):
    """A user-facing problem with the uploaded files or options."""


MONTHS = {
    'jan': 'January', 'feb': 'February', 'mar': 'March', 'apr': 'April',
    'may': 'May', 'jun': 'June', 'jul': 'July', 'aug': 'August',
    'sep': 'September', 'oct': 'October', 'nov': 'November', 'dec': 'December',
}

OUTPUT_COLS = [
    "Sr. No.", "Date of Shipping", "Booking Code", "LR No.",
    "Consignee Name", "Consignee - State", "Consignee - City",
    "Estimated Date of Delivery", "Observation", "Pending Days",
]


def _col(df, *cands):
    """Return a column Series by trying candidate names (case/space-insensitive)."""
    norm = {str(c).strip().lower(): c for c in df.columns}
    for cand in cands:
        key = cand.strip().lower()
        if key in norm:
            return df[norm[key]]
    return pd.Series([None] * len(df), index=df.index)


def _norm_lr(value):
    """Normalise an LR for *exact* matching: trim and drop a trailing '.0'."""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def extract_month_label(*names):
    """Pull the month (and year if present) from the uploaded file names."""
    for nm in names:
        low = str(nm).lower()
        best = None
        for token, full in MONTHS.items():
            idx = low.find(token)
            if idx != -1 and (best is None or idx < best[0]):
                best = (idx, token, full)
        if best:
            idx, token, full = best
            year = ''
            m = re.search(r'(\d{2,4})', low[idx + len(token):])
            if m:
                y = m.group(1)
                year = str(2000 + int(y)) if len(y) == 2 else y
            return full, (f"{full} {year}".strip())
    return None, ''


def load_delayed(file, vehicle_type, min_delay, all_in_transit):
    """Read one master file and return the formatted delayed/in-transit rows."""
    sheet = "2W Report" if vehicle_type == "2W" else "CV Report"
    try:
        xls = pd.ExcelFile(file)
    except Exception as e:
        raise ReportError(f"Couldn't open the {vehicle_type} file as an Excel workbook ({e}).")
    if sheet not in xls.sheet_names:
        raise ReportError(
            f'The {vehicle_type} file has no "{sheet}" sheet — is it the right master report?')

    df = pd.read_excel(xls, sheet_name=sheet)
    df.columns = [str(c) for c in df.columns]

    status = _col(df, "Status").astype(str).str.strip().str.lower()
    days = pd.to_numeric(_col(df, "Days"), errors="coerce")

    in_transit = status == "in transit"
    if all_in_transit:
        mask = in_transit
    else:
        mask = in_transit & (days <= -int(min_delay))

    sel = df[mask].reset_index(drop=True)
    pend = (-days[mask]).reset_index(drop=True)

    out = pd.DataFrame({
        "Sr. No.": list(range(1, len(sel) + 1)),
        "Date of Shipping": pd.to_datetime(_col(sel, "Date of Shipping"), errors="coerce").dt.normalize(),
        "Booking Code": _col(sel, "Booking Code"),
        "LR No.": _col(sel, "Lr No.", "Lr No. ", "LR No.", "LR Number"),
        "Consignee Name": _col(sel, "Consignee Name"),
        "Consignee - State": _col(sel, "Consignee - State", "State"),
        "Consignee - City": _col(sel, "Consignee - City", "City"),
        "Estimated Date of Delivery": pd.to_datetime(
            _col(sel, "Estimated Date of Delivery ( Based upon Agreed TAT)",
                 "Estimated Date of Delivery (Based upon Agreed TAT)", "ETD"),
            errors="coerce").dt.normalize(),
        "Observation": "",
        "Pending Days": pend.astype("Int64"),
    })
    out = out.sort_values("Pending Days", ascending=False, na_position="last").reset_index(drop=True)
    out["Sr. No."] = list(range(1, len(out) + 1))
    return out


# ── Styling ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_data_sheet(ws, row_count):
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = header_align
        cell.border = _BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    widths = {"A": 8, "B": 16, "C": 14, "D": 18, "E": 28,
              "F": 18, "G": 18, "H": 22, "I": 28, "J": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    date_cols = (2, 8)
    light = PatternFill("solid", fgColor="EBF3FB")
    red = PatternFill("solid", fgColor="FFCCCC")
    orange = PatternFill("solid", fgColor="FFE5CC")
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for ridx, row in enumerate(ws.iter_rows(min_row=2, max_row=row_count + 1), start=2):
        pd_cell = row[9]  # Column J = Pending Days
        try:
            dv = int(pd_cell.value) if pd_cell.value is not None else 0
        except (TypeError, ValueError):
            dv = 0
        if dv >= 7:
            rfill = red
        elif dv >= 5:
            rfill = orange
        else:
            rfill = light if ridx % 2 == 0 else None
        for cidx, cell in enumerate(row, start=1):
            cell.font = data_font
            cell.border = _BORDER
            if cidx in date_cols:
                cell.number_format = "DD-MMM-YY"
                cell.alignment = center
            elif cidx in (1, 3, 10):
                cell.alignment = center
            else:
                cell.alignment = left
            if rfill:
                cell.fill = rfill


# ── Orchestrator ─────────────────────────────────────────────────────
def generate(file_2w, file_cv, name_2w, name_cv, min_delay, all_in_transit):
    """Build the pendency workbook (2W + CV sheets). Returns (BytesIO, summary dict).

    The Observation column is left blank here; it is filled afterwards from the
    uploaded observation CSV(s) — see load_observation_csvs / apply_observations.
    """
    out_2w = load_delayed(file_2w, "2W", min_delay, all_in_transit)
    out_cv = load_delayed(file_cv, "CV", min_delay, all_in_transit)
    month_full, month_label = extract_month_label(name_2w, name_cv)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out_2w.to_excel(writer, sheet_name="2W", index=False)
        out_cv.to_excel(writer, sheet_name="CV", index=False)
        _style_data_sheet(writer.sheets["2W"], len(out_2w))
        _style_data_sheet(writer.sheets["CV"], len(out_cv))

    buf.seek(0)
    summary = {
        "count_2w": len(out_2w),
        "count_cv": len(out_cv),
        "total": len(out_2w) + len(out_cv),
        "month": month_label or "—",
    }
    return buf, summary


# ── CSV observation override ─────────────────────────────────────────
_LR_CANDS = ("LR No.", "LR No", "LR Number", "LRN", "Lr No.", "Lr No")
_OBS_CANDS = ("Observation", "Observations", "Status", "Remark", "Remarks", "Comment", "Comments")


def _find_col(df, *cands):
    """Return the actual column name matching a candidate (case/space-insensitive), or None."""
    norm = {str(c).strip().lower(): c for c in df.columns}
    for cand in cands:
        key = cand.strip().lower()
        if key in norm:
            return norm[key]
    return None


def load_observation_csvs(files):
    """Read one or more observation CSVs into a {normalised_lr: observation} map.

    Each CSV must have an LR column (e.g. 'LR No.' / 'LR Number') and an
    Observation column (e.g. 'Observation' / 'Status'). Later rows/files win on
    duplicate LRs. Returns (obs_map, stats). Raises ReportError on bad input.
    """
    obs_map, records, duplicates, file_count = {}, 0, 0, 0
    for f in files:
        name = getattr(f, "name", "CSV")
        try:
            df = pd.read_csv(f, dtype=str, keep_default_na=False)
        except Exception as e:
            raise ReportError(f"Couldn't read observation CSV '{name}' ({e}).")
        df.columns = [str(c) for c in df.columns]

        lr_col = _find_col(df, *_LR_CANDS)
        obs_col = _find_col(df, *_OBS_CANDS)
        if lr_col is None or obs_col is None:
            raise ReportError(
                f"'{name}' needs an LR column (e.g. 'LR No.' / 'LR Number') and an "
                f"Observation column (e.g. 'Observation' / 'Status').")
        file_count += 1

        for lr, obs in zip(df[lr_col], df[obs_col]):
            key = _norm_lr(lr)
            text = str(obs).strip()
            if not key or key.lower() == "nan" or not text:
                continue
            records += 1
            if key in obs_map:
                duplicates += 1
            obs_map[key] = text  # later rows/files override

    if not obs_map:
        raise ReportError("No usable observation rows found in the uploaded CSV(s).")
    return obs_map, {
        "files": file_count, "records": records,
        "duplicates": duplicates, "unique": len(obs_map),
    }


def apply_observations(report_bytes, obs_map):
    """Overlay observations onto an existing pendency workbook by exact LR match.

    Opens the workbook with openpyxl and writes *only* the Observation cells of
    matched rows — every other cell, value, and style is left untouched. Returns
    (BytesIO, stats) where stats includes matched count and the observation LRs
    that matched no report row.
    """
    wb = load_workbook(io.BytesIO(report_bytes))
    matched, report_rows, matched_keys = 0, 0, set()

    for ws in wb.worksheets:
        headers = {str(c.value).strip().lower(): c.column
                   for c in ws[1] if c.value is not None}
        lr_idx = next((headers[c] for c in ("lr no.", "lr no", "lr number", "lrn")
                       if c in headers), None)
        obs_idx = headers.get("observation")
        if lr_idx is None or obs_idx is None:
            continue

        for row in ws.iter_rows(min_row=2):
            lr_cell = row[lr_idx - 1]
            if lr_cell.value is None or str(lr_cell.value).strip() == "":
                continue
            report_rows += 1
            key = _norm_lr(lr_cell.value)
            if key in obs_map:
                row[obs_idx - 1].value = obs_map[key]
                matched += 1
                matched_keys.add(key)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    unmatched = sorted(set(obs_map) - matched_keys)
    stats = {
        "matched": matched, "report_rows": report_rows,
        "unmatched_obs": unmatched, "unmatched_count": len(unmatched),
    }
    return buf, stats
