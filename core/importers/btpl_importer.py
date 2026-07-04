import openpyxl
import datetime
from django.utils import timezone
from core.models import ImportJob, ImportErrorRecord, Shipment, ShipmentStatus, Vehicle, Driver
from core.btpl import get_column_mapping
from core.utils.date_parser import parse_excel_date

class BtplImporter:
    def process_btpl_workbook(self, job_id, file_path):
        """Shadow importer that parses BTPL workbooks and saves to PostgreSQL."""
        job = ImportJob.objects.get(id=job_id)
        job.status = 'RUNNING'
        job.started_at = timezone.now()
        job.save()

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            from core.models import BtplWorkbook
            wb_record = BtplWorkbook.objects.filter(is_active=True).first()
            sheet_name = wb_record.active_sheet if wb_record else 'JUN 26'

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"{sheet_name} not found in BTPL workbook")
            
            sheet = wb[sheet_name]
            mapping = get_column_mapping(sheet)
            
            if not mapping:
                raise ValueError("Failed to map columns for BTPL workbook")

            processed = 0
            failed = 0

            # Start from row 2
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    def get_val(key):
                        col = mapping.get(key)
                        if col:
                            return sheet.cell(row=row_idx, column=col).value
                        return None
                    
                    pickup_date = get_val('pickup_date')
                    lr_number = get_val('lr_number')
                    name = get_val('name')

                    # Skip empty rows (must have at least one key field)
                    if not pickup_date and not lr_number and not name:
                        continue

                    # Process row
                    self._process_btpl_row(job, row_idx, get_val)
                    processed += 1
                except Exception as e:
                    failed += 1
                    ImportErrorRecord.objects.create(
                        import_job=job,
                        row_number=row_idx,
                        error_message=str(e)
                    )

            job.processed_rows = processed
            job.failed_rows = failed
            job.total_rows = processed + failed
            
            if failed > 0 and processed > 0:
                job.status = 'PARTIAL_SUCCESS'
            elif failed > 0:
                job.status = 'FAILED'
            else:
                job.status = 'COMPLETED'

            job.completed_at = timezone.now()
            job.save()
            wb.close()

        except Exception as ex:
            job.status = 'FAILED'
            job.completed_at = timezone.now()
            job.save()
            ImportErrorRecord.objects.create(
                import_job=job,
                row_number=0,
                error_message=f"Fatal Error: {str(ex)}"
            )

    def _process_btpl_row(self, job, row_idx, get_val):
        from django.db import transaction
        with transaction.atomic():
            pickup_date = get_val('pickup_date')
            
            pickup_date = parse_excel_date(pickup_date)

            lr = str(get_val('lr_number') or '')
            source_key = f"{lr}|{pickup_date.isoformat() if pickup_date else ''}"

            if not lr and not pickup_date:
                raise ValueError("Cannot import row: missing both lr_number and pickup_date")

            # Create Shipment
            shipment, _ = Shipment.objects.update_or_create(
                shipment_type='BTPL',
                source_key=source_key,
                defaults={
                    'dispatch_date': pickup_date,
                    'metadata': {
                        'lr_number': lr,
                        'name': str(get_val('name') or ''),
                        'address': str(get_val('address') or ''),
                        'contact_person': str(get_val('contact_person') or ''),
                        'contact_number': str(get_val('contact_number') or ''),
                        'city': str(get_val('city') or ''),
                        'state': str(get_val('state') or ''),
                        'boxes': get_val('boxes'),
                        'weight_ef': get_val('weight_ef'),
                        'weight_opt': get_val('weight_opt'),
                        'rate': get_val('rate'),
                        'amount': get_val('amount'),
                        'vendor': str(get_val('vendor') or ''),
                        'vendor_rate': get_val('vendor_rate'),
                        'vendor_payment': get_val('vendor_payment')
                    }
                }
            )

            # BTPL does not explicitly have ETD/Delivery dates in the default header map,
            # but we can set a draft status.
            ShipmentStatus.objects.get_or_create(
                shipment=shipment,
                status='DRAFT'
            )
