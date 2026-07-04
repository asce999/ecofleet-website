import openpyxl
import datetime
from django.utils import timezone
from core.models import ImportJob, ImportErrorRecord, Shipment, ShipmentStatus, Vehicle, Driver
from core.ftl import get_column_mapping
from core.utils.date_parser import parse_excel_date, parse_excel_datetime

class ExcelImporter:
    def process_ftl_workbook(self, job_id, file_path):
        """Shadow importer that parses FTL workbooks and saves to PostgreSQL."""
        job = ImportJob.objects.get(id=job_id)
        job.status = 'RUNNING'
        job.started_at = timezone.now()
        job.save()

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if 'Sheet1' not in wb.sheetnames:
                raise ValueError("Sheet1 not found in FTL workbook")
            
            sheet = wb['Sheet1']
            mapping = get_column_mapping(sheet)
            
            if not mapping:
                raise ValueError("Failed to map columns for FTL workbook")

            processed = 0
            failed = 0

            # Start from row 2
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    def get_val(key):
                        col = mapping.get(key)
                        if col:
                            return sheet.cell(row=row_idx, column=col).value
                        return None
                    
                    booking_date = get_val('booking_date')
                    lr_number = get_val('lr_number')
                    vehicle_number = get_val('vehicle_number')

                    # Skip empty rows (must have at least one key field)
                    if not booking_date and not lr_number and not vehicle_number:
                        continue

                    # Process row
                    self._process_ftl_row(job, row_idx, get_val)
                    processed += 1
                except Exception as e:
                    failed += 1
                    ImportErrorRecord.objects.create(
                        import_job=job,
                        row_number=row_idx,
                        error_message=str(e)
                    )

            job.processed_rows = processed
            job.failed_rows = failed
            job.total_rows = processed + failed
            
            if failed > 0 and processed > 0:
                job.status = 'PARTIAL_SUCCESS'
            elif failed > 0:
                job.status = 'FAILED'
            else:
                job.status = 'COMPLETED'

            job.completed_at = timezone.now()
            job.save()
            wb.close()

        except Exception as ex:
            job.status = 'FAILED'
            job.completed_at = timezone.now()
            job.save()
            ImportErrorRecord.objects.create(
                import_job=job,
                row_number=0,
                error_message=f"Fatal Error: {str(ex)}"
            )

    def _process_ftl_row(self, job, row_idx, get_val):
        from django.db import transaction
        with transaction.atomic():
            # Get Vehicle
            vehicle_number = get_val('vehicle_number')
            vehicle_obj = None
            if vehicle_number:
                vehicle_obj, _ = Vehicle.objects.get_or_create(
                    registration_number=str(vehicle_number).strip().upper()
                )

            # Map fields
            origin = get_val('origin') or ""
            destination = get_val('destination') or ""
            booking_date = get_val('booking_date')
            
            booking_date = parse_excel_date(booking_date)

            etd = get_val('etd')
            delivery_date = get_val('delivery_date')
            expected_eta = parse_excel_datetime(etd)
            actual_eta = parse_excel_datetime(delivery_date)

            lr = str(get_val('lr_number') or '')
            source_key = f"{lr}|{booking_date.isoformat() if booking_date else ''}"

            if not lr and not booking_date:
                raise ValueError("Cannot import row: missing both lr_number and booking_date")

            # Create Shipment
            shipment, _ = Shipment.objects.update_or_create(
                shipment_type='FTL',
                source_key=source_key,
                defaults={
                    'origin': str(origin),
                    'destination': str(destination),
                    'dispatch_date': booking_date,
                    'expected_eta': expected_eta,
                    'actual_eta': actual_eta,
                    'vehicle': vehicle_obj,
                    'metadata': {
                        'lr_number': lr,
                        'consignor': str(get_val('consignor')) if get_val('consignor') else "",
                        'consignee': str(get_val('consignee')) if get_val('consignee') else "",
                        'vendor': str(get_val('vendor')) if get_val('vendor') else "",
                    }
                }
            )

            # Determine Status
            status_val = 'DRAFT'
            if delivery_date:
                status_val = 'DELIVERED'
            elif etd:
                status_val = 'IN_TRANSIT'

            # get_or_create (not create) so re-importing the same row does not
            # bloat the status log with duplicate rows (F-02 idempotency).
            ShipmentStatus.objects.get_or_create(
                shipment=shipment,
                status=status_val
            )
