from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from .models import Pincode
from django.template import loader
import datetime
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect
from django.utils.http import url_has_allowed_host_and_scheme
from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone
import os
import io
import pandas as pd
from django.core.files.base import ContentFile
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404
from django.views.decorators.cache import never_cache

from .forms import CofForm, WorkbookUploadForm, PendencyForm, MorningForm, PrevMonthUpdateForm, BtplShipmentForm, BtplWorkbookUploadForm, AttendanceWorkbookUploadForm
from . import cof, pendency, morning, prev_month, btpl, attendance

from .decorators import staff_required, tool_permission_required, director_required
from .models import ToolRun, ToolRunFile, CofWorkbook, BtplWorkbook, AttendanceWorkbook

def home(request):
    return render(request, 'core/home.html')

def services(request):
    return render(request, 'core/services.html')

def contact(request):
    return render(request, 'core/contact.html')

def about(request):
    return render(request, 'core/about.html')

def privacy(request):
    return render(request, 'core/privacy.html')

def sitemap(request):
    xml_content = '''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url>
    <loc>http://ecofleetexpress.com/</loc>
    <lastmod>2026-06-07</lastmod>
    <changefreq>monthly</changefreq>
    <priority>1.0</priority>
  </url>
  <url>
    <loc>http://ecofleetexpress.com/about/</loc>
    <lastmod>2026-06-07</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>http://ecofleetexpress.com/services/</loc>
    <lastmod>2026-06-07</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.9</priority>
  </url>
  <url>
    <loc>http://ecofleetexpress.com/contact/</loc>
    <lastmod>2026-06-07</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.7</priority>
  </url>
  <url>
    <loc>http://ecofleetexpress.com/privacy/</loc>
    <lastmod>2026-06-07</lastmod>
    <changefreq>yearly</changefreq>
    <priority>0.3</priority>
  </url>
</urlset>'''
    return HttpResponse(xml_content, content_type='application/xml')

def find_location(request):
    result = None
    pincode = request.GET.get('pincode', '').strip()

    if pincode:
        try:
            pin_obj = Pincode.objects.get(pin=pincode)
            result = {
                'found': True,
                'pin': pin_obj.pin,
                'city': pin_obj.city,
                'state': pin_obj.state,
                'location_type': pin_obj.location_type,
            }
        except Pincode.DoesNotExist:
            result = {'found': False}

    return render(request, 'core/find_location.html', {'result': result, 'pincode': pincode})

# ─────────────────────────────────────────────
# EMPLOYEE PORTAL
# ─────────────────────────────────────────────
def portal_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('dashboard')

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            user = form.get_user()
            if not user.is_staff:
                messages.error(request, "This account isn't authorised for the employee portal.")
            else:
                auth_login(request, user)
                nxt = request.POST.get('next') or request.GET.get('next')
                if nxt and url_has_allowed_host_and_scheme(nxt, allowed_hosts={request.get_host()}):
                    return redirect(nxt)
                return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'core/portal/login.html', {'form': form})


def portal_logout(request):
    auth_logout(request)
    messages.success(request, "You've been logged out.")
    return redirect('portal_login')


@staff_required
def dashboard(request):
    User = get_user_model()
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ── Pincode coverage ──
    total_pincodes = Pincode.objects.count()
    oda = Pincode.objects.filter(location_type__iexact='ODA').count()
    non_oda = total_pincodes - oda

    # ── Operational counts (light up as the tools come online) ──
    cofs_month = ToolRun.objects.filter(
        tool=ToolRun.TOOL_COF, status=ToolRun.STATUS_SUCCESS,
        created_at__gte=month_start).count()
    reports_month = ToolRun.objects.filter(
        tool__in=[ToolRun.TOOL_MORNING, ToolRun.TOOL_PENDENCY, ToolRun.TOOL_PREV_MONTH],
        status=ToolRun.STATUS_SUCCESS, created_at__gte=month_start).count()
    total_runs = ToolRun.objects.count()
    team_members = User.objects.filter(is_staff=True, is_active=True).count()

    # ── Top states by coverage ──
    top_states = list(
        Pincode.objects.values('state')
        .annotate(c=Count('id')).order_by('-c')[:8])
    state_labels = [s['state'] or 'Unknown' for s in top_states]
    state_data = [s['c'] for s in top_states]

    # ── Tool activity, last 14 days ──
    since = (now - datetime.timedelta(days=13)).date()
    by_day = {
        r['d']: r['c'] for r in
        ToolRun.objects.filter(created_at__date__gte=since)
        .annotate(d=TruncDate('created_at')).values('d')
        .annotate(c=Count('id'))
    }
    activity_labels, activity_data = [], []
    for i in range(13, -1, -1):
        day = (now - datetime.timedelta(days=i)).date()
        activity_labels.append(day.strftime('%d %b'))
        activity_data.append(by_day.get(day, 0))

    # ── Tool usage breakdown ──
    usage = {r['tool']: r['c'] for r in
            ToolRun.objects.values('tool').annotate(c=Count('id'))}
    tool_counts = [usage.get(ToolRun.TOOL_COF, 0),
                    usage.get(ToolRun.TOOL_MORNING, 0),
                    usage.get(ToolRun.TOOL_PENDENCY, 0),
                    usage.get(ToolRun.TOOL_PREV_MONTH, 0)]

    chart_data = {
        'coverage': {'oda': oda, 'non_oda': non_oda},
        'states': {'labels': state_labels, 'data': state_data},
        'activity': {'labels': activity_labels, 'data': activity_data},
        'tools': {'labels': ['COF', 'Morning Report', 'Pendency', 'Previous Month Update'], 'data': tool_counts},
    }

    ctx = {
        'active': 'dashboard',
        'today': datetime.date.today().strftime('%A, %d %B %Y'),
        'total_pincodes': total_pincodes,
        'oda': oda,
        'non_oda': non_oda,
        'states_served': Pincode.objects.values('state').distinct().count(),
        'cofs_month': cofs_month,
        'reports_month': reports_month,
        'total_runs': total_runs,
        'team_members': team_members,
        'chart_data': chart_data,
        'recent_runs': ToolRun.objects.select_related('user')[:8],
    }
    return render(request, 'core/portal/dashboard.html', ctx)


# ── COF Generator (upload-once server-workbook model) ──
@staff_required
@tool_permission_required('cof')
def cof_generator(request):
    wb_obj = CofWorkbook.active()
    if not wb_obj:
        return redirect('cof_workbook')

    wb_path = wb_obj.file.path

    try:
        preview = cof.get_next_cof_info(wb_path)
        preview_error = None
    except Exception as e:
        preview, preview_error = None, str(e)

    if request.method == 'POST':
        form = CofForm(request.POST)
        if form.is_valid():
            data = form.to_cof_data()
            try:
                result = cof.generate_cof(data, wb_path)
            except (cof.COFLockTimeout, cof.WorkbookInUse,
                    cof.WorkbookInvalid, cof.AssetMissing) as e:
                messages.error(request, str(e))
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_COF,
                    status=ToolRun.STATUS_FAILED,
                    reference=preview['cof_number'] if preview else '',
                    detail=f"Error: {e}")
                messages.error(request, f"Failed to generate COF: {e}")
            else:
                run = ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_COF,
                    status=ToolRun.STATUS_SUCCESS, reference=result['cof_number'],
                    detail=f"Sheet: {result['sheet_name']} · Consignee: {data['consignee_name']}")
                ToolRunFile.objects.create(
                    run=run, label="COF document", download_name=result['docx_name'],
                    file=ContentFile(result['docx'].getvalue(), name=result['docx_name']))
                messages.success(request, f"{result['cof_number']} generated successfully.")
                return redirect('cof_success', pk=run.pk)
        else:
            messages.error(request, "Please fix the highlighted fields.")
    else:
        form = CofForm()

    return render(request, 'core/portal/cof_form.html', {
        'active': 'cof', 'form': form, 'preview': preview,
        'preview_error': preview_error, 'wb': wb_obj,
    })


@staff_required
@tool_permission_required('cof')
def cof_workbook(request):
    """Upload / replace / remove the active tracking workbook."""
    wb_obj = CofWorkbook.active()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'remove' and wb_obj:
            wb_obj.is_active = False
            wb_obj.save(update_fields=['is_active'])
            messages.success(request, "Active workbook removed. Upload one to continue.")
            return redirect('cof_workbook')

        form = WorkbookUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.cleaned_data['workbook']
            new = CofWorkbook.objects.create(
                file=upload, original_name=upload.name,
                uploaded_by=request.user, is_active=False)
            try:
                cof.validate_workbook(new.file.path)
            except cof.WorkbookInvalid as e:
                new.file.delete(save=False)
                new.delete()
                messages.error(request, str(e))
                return redirect('cof_workbook')
            CofWorkbook.objects.filter(is_active=True).update(is_active=False)
            new.is_active = True
            new.save(update_fields=['is_active'])
            messages.success(request, f"“{upload.name}” is now the active tracking workbook.")
            return redirect('cof_generator')
        else:
            messages.error(request, "Please choose a valid .xlsx file.")
    else:
        form = WorkbookUploadForm()

    info = None
    if wb_obj:
        try:
            info = cof.get_next_cof_info(wb_obj.file.path)
        except Exception:
            info = None
    return render(request, 'core/portal/cof_workbook.html', {
        'active': 'cof', 'form': form, 'wb': wb_obj, 'info': info,
    })


@staff_required
@tool_permission_required('cof')
def cof_success(request, pk):
    run = get_object_or_404(ToolRun.objects.prefetch_related('files'),
                            pk=pk, tool=ToolRun.TOOL_COF, status=ToolRun.STATUS_SUCCESS)
    return render(request, 'core/portal/cof_success.html', {'active': 'cof', 'run': run})


@staff_required
@tool_permission_required('cof')
def cof_history(request):
    wb_obj = CofWorkbook.active()
    q = request.GET.get('q', '').strip()
    rows = []
    if wb_obj:
        rows = list(reversed(cof.load_history(wb_obj.file.path)))
        if q:
            ql = q.lower()
            rows = [r for r in rows if any(ql in str(v).lower() for v in r.values())]
    display = [{
        'num': r.get('#', ''), 'lr': r.get('LR Number', ''),
        'invoice': r.get('Invoice Number', ''), 'dealer': r.get('Dealer', ''),
        'state': r.get('State', ''), 'claim': r.get('Claim Amount', ''),
        'status': r.get('Status Delhivery', ''), 'cof_date': r.get('COF Date', ''),
        'optlog': r.get('Status Optlog', ''),
    } for r in rows]
    return render(request, 'core/portal/cof_history.html', {
        'active': 'cof', 'rows': display, 'q': q, 'total': len(display), 'wb': wb_obj,
    })


@staff_required
@tool_permission_required('cof')
def cof_workbook_download(request):
    wb_obj = CofWorkbook.active()
    if not wb_obj:
        raise Http404("No active workbook.")
    return FileResponse(wb_obj.file.open('rb'), as_attachment=True,
                        filename=wb_obj.original_name)


@staff_required
def download_file(request, file_id):
    f = get_object_or_404(ToolRunFile, pk=file_id)
    if not f.file:
        raise Http404("File not available.")

    from core.models import UserProfile
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    tool_map = {
        'COF': 'cof',
        'MORNING': 'morning',
        'PENDENCY': 'pendency',
        'PREV_MONTH': 'prev_month',
        'BTPL': 'btpl',
        'ATTENDANCE': 'attendance',
    }
    required_perm = tool_map.get(f.run.tool)
    if required_perm and not getattr(profile, f"can_use_{required_perm}", False):
        messages.error(request, f"You do not have permission to download files generated by the {f.run.get_tool_display()} tool.")
        return redirect('dashboard')

    return FileResponse(f.file.open('rb'), as_attachment=True,
                        filename=f.download_name or os.path.basename(f.file.name))


@staff_required
@tool_permission_required('morning')
def morning_report(request):
    if request.method == 'POST':
        form = MorningForm(request.POST, request.FILES)
        delhivery_files = request.FILES.getlist('delhivery_files')
        ok = form.is_valid()
        if not delhivery_files:
            messages.error(request, "Upload at least one Delhivery CSV.")
            ok = False
        elif any(not f.name.lower().endswith('.csv') for f in delhivery_files):
            messages.error(request, "Delhivery files must be .csv.")
            ok = False

        if ok:
            try:
                result = morning.generate(
                    delhivery_files, form.cleaned_data['file_2w'], form.cleaned_data['file_cv'])
            except morning.ReportError as e:
                messages.error(request, str(e))
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_MORNING,
                    status=ToolRun.STATUS_FAILED, detail=f"Error: {e}")
                messages.error(request, f"Failed to generate report: {e}")
            else:
                s = result['summary']
                run = ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_MORNING,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"Morning Report {datetime.date.today():%d %b %Y}",
                    detail=(f"2W: {s['two_w']['updated']} updated / {s['two_w']['new']} new · "
                            f"CV: {s['cv']['updated']} updated / {s['cv']['new']} new · "
                            f"{s['delhivery_rows']} Delhivery rows"))
                for key in ("2W", "CV"):
                    buf, fname = result[key]
                    ToolRunFile.objects.create(
                        run=run, label=f"{key} master", download_name=fname,
                        file=ContentFile(buf.getvalue(), name=fname))
                messages.success(request, "Morning report generated.")
                return redirect('tool_result', pk=run.pk)
        # fall through to re-render with errors
    else:
        form = MorningForm()
    return render(request, 'core/portal/morning_form.html',
                  {'active': 'morning', 'form': form})


@staff_required
@tool_permission_required('pendency')
def pendency_report(request):
    latest_run = ToolRun.objects.filter(tool=ToolRun.TOOL_PENDENCY, status=ToolRun.STATUS_SUCCESS).first()
    latest_file = latest_run.files.last() if latest_run else None
    
    preview = None
    all_lrs = {}
    if latest_file:
        preview = _pendency_preview(latest_file)
        all_lrs = _get_all_lrs_from_workbook(latest_file)

    if request.method == 'POST':
        form = PendencyForm(request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                buf, summary = pendency.generate(
                    cd['file_2w'], cd['file_cv'],
                    cd['file_2w'].name, cd['file_cv'].name,
                    cd.get('min_delay_days') or 1,
                    cd.get('all_in_transit') or False)
            except pendency.ReportError as e:
                messages.error(request, str(e))
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_PENDENCY,
                    status=ToolRun.STATUS_FAILED, detail=f"Error: {e}")
                messages.error(request, f"Failed to generate report: {e}")
            else:
                fname = f"{cd['report_name']}.xlsx"
                run = ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_PENDENCY,
                    status=ToolRun.STATUS_SUCCESS, reference=cd['report_name'],
                    detail=(f"2W: {summary['count_2w']} · CV: {summary['count_cv']} "
                            f"shipments · Month: {summary['month']}"))
                ToolRunFile.objects.create(
                    run=run, label="Pendency report", download_name=fname,
                    file=ContentFile(buf.getvalue(), name=fname))
                messages.success(request, "Pendency report generated.")
                return redirect('tool_result', pk=run.pk)
        else:
            messages.error(request, "Please fix the highlighted fields.")
    else:
        form = PendencyForm()
    
    return render(request, 'core/portal/pendency_form.html', {
        'active': 'pendency',
        'form': form,
        'latest_run': latest_run,
        'current': latest_file,
        'preview': preview,
        'all_lrs': all_lrs,
    })


def _read_file_bytes(tool_file):
    """Read a ToolRunFile's bytes (open/close safely)."""
    f = tool_file.file
    f.open('rb')
    try:
        return f.read()
    finally:
        f.close()


def _pendency_preview(tool_file, max_rows=15):
    """Build a small per-sheet preview of a pendency workbook for the UI."""
    try:
        data = _read_file_bytes(tool_file)
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return []
    sheets = []
    for name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=name, nrows=max_rows).fillna("")
        except Exception:
            continue
        sheets.append({
            'name': name,
            'columns': [str(c) for c in df.columns],
            'rows': df.astype(str).values.tolist(),
        })
    return sheets


def _get_all_lrs_from_workbook(tool_file):
    """Get all LR numbers from the 2W and CV sheets as a dict: {sheet_name: [lr1, lr2, ...]}."""
    try:
        data = _read_file_bytes(tool_file)
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return {}
    
    result = {}
    for name in xls.sheet_names:
        if name in ("2W", "CV"):
            try:
                df = pd.read_excel(xls, sheet_name=name)
                df.columns = [str(c).strip().lower() for c in df.columns]
                lr_col = next((c for c in df.columns if c in ("lr no.", "lr no", "lr number", "lrn")), None)
                if lr_col:
                    lrs = df[lr_col].dropna().astype(str).str.strip().str.replace(r'\.0$', '', regex=True).tolist()
                    result[name] = lrs
            except Exception:
                continue
    return result


@staff_required
@tool_permission_required('pendency')
def pendency_observations(request, pk):
    """Override / fill the Observation column of an existing pendency report by
    uploading one or more observation CSVs, matched strictly by LR Number."""
    run = get_object_or_404(
        ToolRun.objects.prefetch_related('files'),
        pk=pk, tool=ToolRun.TOOL_PENDENCY, status=ToolRun.STATUS_SUCCESS)
    current = run.files.last()  # newest report (original, or a prior override)
    if not current or not current.file:
        raise Http404("No pendency report file to enrich.")

    if request.method == 'POST':
        obs_files = request.FILES.getlist('observation_files')
        if not obs_files:
            messages.error(request, "Upload at least one observation CSV.")
        elif any(not f.name.lower().endswith('.csv') for f in obs_files):
            messages.error(request, "Observation files must be .csv.")
        else:
            try:
                obs_map, load_stats = pendency.load_observation_csvs(obs_files)
                report_bytes = _read_file_bytes(current)
                buf, apply_stats = pendency.apply_observations(report_bytes, obs_map)
            except pendency.ReportError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Failed to apply observations: {e}")
            else:
                fname = f"{run.reference or 'pendency'}_with_observations.xlsx"
                ToolRunFile.objects.create(
                    run=run, label="Pendency report (observations updated)",
                    download_name=fname,
                    file=ContentFile(buf.getvalue(), name=fname))
                extra = (f"Observations applied: {apply_stats['matched']} matched "
                         f"from {load_stats['unique']} CSV record(s)")
                if apply_stats['unmatched_count']:
                    extra += f" · {apply_stats['unmatched_count']} CSV LR(s) unmatched"
                run.detail = f"{run.detail} · {extra}"
                run.save(update_fields=['detail'])
                messages.success(
                    request, f"{apply_stats['matched']} observation(s) applied by exact LR match.")
                if apply_stats['unmatched_count']:
                    sample = ", ".join(apply_stats['unmatched_obs'][:10])
                    more = "…" if apply_stats['unmatched_count'] > 10 else ""
                    messages.warning(
                        request, f"{apply_stats['unmatched_count']} LR(s) from the CSV(s) "
                                 f"matched no report row: {sample}{more}")
                return redirect('tool_result', pk=run.pk)

    return render(request, 'core/portal/pendency_observations.html', {
        'active': 'pendency', 'run': run, 'current': current,
        'preview': _pendency_preview(current),
        'all_lrs': _get_all_lrs_from_workbook(current),
    })


@staff_required
def tool_result(request, pk):
    run = get_object_or_404(ToolRun.objects.prefetch_related('files'),
                            pk=pk, status=ToolRun.STATUS_SUCCESS)
    active = {'COF': 'cof', 'MORNING': 'morning', 'PENDENCY': 'pendency', 'PREV_MONTH': 'prev_month'}.get(run.tool, '')

    # Check permissions
    from core.models import UserProfile
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if active and not getattr(profile, f"can_use_{active}", False):
        messages.error(request, f"You do not have permission to view results for the {run.get_tool_display()} tool.")
        return redirect('dashboard')

    return render(request, 'core/portal/tool_result.html', {'active': active, 'run': run})


@staff_required
@tool_permission_required('prev_month')
def prev_month_update(request):
    if request.method == 'POST':
        form = PrevMonthUpdateForm(request.POST, request.FILES)
        delhivery_files = request.FILES.getlist('delhivery_files')
        ok = form.is_valid()
        if not delhivery_files:
            messages.error(request, "Upload at least one Delhivery CSV.")
            ok = False
        elif any(not f.name.lower().endswith('.csv') for f in delhivery_files):
            messages.error(request, "Delhivery files must be .csv.")
            ok = False

        if ok:
            try:
                result = prev_month.generate(
                    delhivery_files, form.cleaned_data['file_2w'], form.cleaned_data['file_cv'])
            except prev_month.ReportError as e:
                messages.error(request, str(e))
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_PREV_MONTH,
                    status=ToolRun.STATUS_FAILED, detail=f"Error: {e}")
                messages.error(request, f"Failed to run previous month update: {e}")
            else:
                s = result['summary']
                run = ToolRun.objects.create(
                    user=request.user, tool=ToolRun.TOOL_PREV_MONTH,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"Previous Month Update {datetime.date.today():%d %b %Y}",
                    detail=(f"2W: {s['two_w']['updated']} updated ({s['two_w']['delivered']} delivered / {s['two_w']['in_transit']} in transit) · "
                            f"CV: {s['cv']['updated']} updated ({s['cv']['delivered']} delivered / {s['cv']['in_transit']} in transit) · "
                            f"{s['delhivery_rows']} Delhivery rows"))
                for key in ("2W", "CV"):
                    buf, fname = result[key]
                    ToolRunFile.objects.create(
                        run=run, label=f"{key} master", download_name=fname,
                        file=ContentFile(buf.getvalue(), name=fname))
                messages.success(request, "Previous month update completed.")
                return redirect('tool_result', pk=run.pk)
    else:
        form = PrevMonthUpdateForm()
    return render(request, 'core/portal/prev_month_form.html',
                  {'active': 'prev_month', 'form': form})


@staff_required
@director_required
def portal_users(request):
    from django.contrib.auth.models import User
    from core.models import UserProfile

    users = User.objects.filter(is_staff=True).order_by('username')
    for u in users:
        UserProfile.objects.get_or_create(user=u)

    if request.method == 'POST':
        for u in users:
            if u == request.user:
                continue

            profile = u.profile
            role = request.POST.get(f"role_{u.id}", "Employee")
            profile.role = role

            profile.can_use_cof = f"cof_{u.id}" in request.POST
            profile.can_use_morning = f"morning_{u.id}" in request.POST
            profile.can_use_pendency = f"pendency_{u.id}" in request.POST
            profile.can_use_prev_month = f"prev_month_{u.id}" in request.POST
            profile.can_use_btpl = f"btpl_{u.id}" in request.POST
            profile.can_use_attendance = f"attendance_{u.id}" in request.POST
            profile.save()

        messages.success(request, "User roles and permissions updated successfully.")
        return redirect('portal_users')

    profiles = [u.profile for u in users]
    return render(request, 'core/portal/users.html', {
        'active': 'users',
        'profiles': profiles,
    })


def get_active_btpl_workbook():
    from .models import BtplWorkbook
    # If there are BtplWorkbook records but no active one, it means sheet was removed entirely!
    wb_count = BtplWorkbook.objects.count()
    if wb_count > 0:
        wb_obj = BtplWorkbook.active()
        if not wb_obj:
            return None, None, None
        return wb_obj, wb_obj.file.path, wb_obj.active_sheet
    else:
        # Initial fresh state: fall back to default
        file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
        sheet_name = 'JUN 26'
        return None, file_path, sheet_name


@staff_required
@tool_permission_required('btpl')
@never_cache
def btpl_sheet(request):
    from django.urls import reverse
    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    
    if not file_path:
        return render(request, 'core/portal/btpl_form.html', {
            'active': 'btpl',
            'no_sheet': True,
        })

    page = int(request.GET.get('page', 1))
    page_data = btpl.get_btpl_page_data(
        file_path, sheet_name=sheet_name, page=page, page_size=20
    )

    if page_data is None:
        return render(request, 'core/portal/btpl_form.html', {
            'active': 'btpl',
            'no_sheet': True,
        })

    totals_row = page_data['totals_row']
    next_row = page_data['next_row']

    context = {
        'active': 'btpl',
        'preview': page_data['preview'],
        'next_row': next_row,
        'totals_row': totals_row,
        'max_shipment_row': totals_row - 1,
        'sheet_name': sheet_name,
        'wb': wb_obj,
    }
    return render(request, 'core/portal/btpl_form.html', context)


@staff_required
@tool_permission_required('btpl')
@never_cache
def btpl_api(request):
    """JSON API endpoint for AJAX BTPL operations."""
    import json
    from django.http import JsonResponse

    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    if not file_path:
        return JsonResponse({'error': 'No active workbook'}, status=404)

    action = request.GET.get('action') or request.POST.get('action', '')

    # GET: fetch row data for editing
    if action == 'get_row':
        row_num = int(request.GET.get('row', 0))
        if row_num < 2:
            return JsonResponse({'error': 'Invalid row'}, status=400)
        row_data = btpl.get_btpl_row_values(file_path, row_num, sheet_name=sheet_name)
        # Convert dates to strings for JSON
        for key in ['pickup_date', 'delivered_on']:
            val = row_data.get(key)
            if isinstance(val, (datetime.datetime, datetime.date)):
                row_data[key] = val.strftime('%Y-%m-%d')
        return JsonResponse({'row_data': row_data})

    # GET: paginated preview
    if action == 'preview':
        page = int(request.GET.get('page', 1))
        preview = btpl.get_btpl_preview(file_path, sheet_name=sheet_name, page=page, page_size=20)
        return JsonResponse({'preview': preview})

    # POST: save row
    if action == 'save' and request.method == 'POST':
        form = BtplShipmentForm(request.POST)
        if form.is_valid():
            row_data = form.cleaned_data
            target_row = row_data['row_num']
            try:
                btpl.add_btpl_shipment(file_path, row_data, sheet_name=sheet_name)
                run = ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_BTPL,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"LR: {row_data.get('lr_number') or ''}",
                    detail=f"Row: {target_row} · Sheet: {sheet_name} · Consignee: {row_data.get('name') or ''}"
                )
                return JsonResponse({'success': True, 'run_id': run.pk, 'row': target_row})
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_BTPL,
                    status=ToolRun.STATUS_FAILED,
                    detail=f"Error writing row {target_row}: {e}"
                )
                return JsonResponse({'error': str(e)}, status=500)
        else:
            errors = {k: v[0] for k, v in form.errors.items()}
            return JsonResponse({'error': 'Validation failed', 'field_errors': errors}, status=400)

    # POST: delete row
    if action == 'delete' and request.method == 'POST':
        row_num = int(request.POST.get('row', 0))
        # Get totals_row to validate
        page_data = btpl.get_btpl_page_data(file_path, sheet_name=sheet_name, page=1, page_size=1)
        totals_row = page_data['totals_row'] if page_data else 64
        if row_num < 2 or row_num >= totals_row:
            return JsonResponse({'error': 'Invalid row number'}, status=400)
        try:
            btpl.clear_btpl_row(file_path, row_num, sheet_name=sheet_name)
            ToolRun.objects.create(
                user=request.user,
                tool=ToolRun.TOOL_BTPL,
                status=ToolRun.STATUS_SUCCESS,
                reference=f"Clear Row: {row_num}",
                detail=f"Cleared all shipment details in Row {row_num} on sheet {sheet_name}"
            )
            return JsonResponse({'success': True, 'row': row_num})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # GET: next available row
    if action == 'next_row':
        next_row = btpl.find_next_btpl_row(file_path, sheet_name=sheet_name)
        return JsonResponse({'next_row': next_row})

    return JsonResponse({'error': 'Unknown action'}, status=400)


@staff_required
@tool_permission_required('btpl')
def btpl_download(request):
    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    if not file_path or not os.path.exists(file_path):
        raise Http404("BTPL Shipments file not found.")
    
    filename = wb_obj.original_name if wb_obj else "BTPL_Shipments.xlsx"
    response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_required
@tool_permission_required('btpl')
def btpl_settings(request):
    """Manage active BTPL shipment workbook and active sheet tab."""
    wb_obj, file_path, current_sheet_name = get_active_btpl_workbook()
    
    # Read sheets from Excel if file_path exists
    sheets = []
    if file_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
        except Exception:
            sheets = ['JUN 26']
    else:
        sheets = ['JUN 26']
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'remove':
            BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
            if BtplWorkbook.objects.count() == 0:
                BtplWorkbook.objects.create(
                    original_name='BTPL_Shipments.xlsx',
                    active_sheet='JUN 26',
                    uploaded_by=request.user,
                    is_active=False
                )
            messages.success(request, "BTPL shipment sheet removed entirely. Portal is now in empty state.")
            return redirect('btpl_settings')
            
        elif action == 'load_default':
            import shutil
            root_file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
            target_dir = os.path.join(settings.MEDIA_ROOT, 'btpl')
            os.makedirs(target_dir, exist_ok=True)
            dest_path = os.path.join(target_dir, 'BTPL_Shipments.xlsx')
            shutil.copy2(root_file_path, dest_path)
            
            # Deactivate any existing
            BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
            
            wb_obj = BtplWorkbook.objects.create(
                original_name='BTPL_Shipments.xlsx',
                active_sheet='JUN 26',
                uploaded_by=request.user,
                is_active=True
            )
            wb_obj.file.name = 'btpl/BTPL_Shipments.xlsx'
            wb_obj.save(update_fields=['file'])
            
            messages.success(request, "Default BTPL shipment workbook loaded successfully.")
            return redirect('btpl_settings')
            
        elif action == 'change_sheet':
            form = BtplWorkbookUploadForm(request.POST, sheets=sheets)
            if form.is_valid():
                sheet_sel = form.cleaned_data.get('active_sheet')
                if wb_obj:
                    wb_obj.active_sheet = sheet_sel
                    wb_obj.save(update_fields=['active_sheet'])
                else:
                    # Copy root file to media directory and register in database
                    import shutil
                    root_file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
                    target_dir = os.path.join(settings.MEDIA_ROOT, 'btpl')
                    os.makedirs(target_dir, exist_ok=True)
                    dest_path = os.path.join(target_dir, 'BTPL_Shipments.xlsx')
                    shutil.copy2(root_file_path, dest_path)
                    
                    wb_obj = BtplWorkbook.objects.create(
                        original_name='BTPL_Shipments.xlsx',
                        active_sheet=sheet_sel,
                        uploaded_by=request.user,
                        is_active=True
                    )
                    wb_obj.file.name = 'btpl/BTPL_Shipments.xlsx'
                    wb_obj.save(update_fields=['file'])
                    
                messages.success(request, f"Active sheet tab changed to '{sheet_sel}'.")
                return redirect('btpl_sheet')
                
        elif action == 'upload':
            form = BtplWorkbookUploadForm(request.POST, request.FILES, sheets=sheets)
            if form.is_valid():
                upload = request.FILES.get('workbook')
                if upload:
                    # Create new workbook record
                    new_wb = BtplWorkbook.objects.create(
                        file=upload, original_name=upload.name,
                        uploaded_by=request.user, is_active=False
                    )
                    
                    # Read sheets from the uploaded workbook to set default active sheet
                    try:
                        import openpyxl
                        temp_wb = openpyxl.load_workbook(new_wb.file.path, read_only=True)
                        uploaded_sheets = temp_wb.sheetnames
                        default_sheet = uploaded_sheets[0] if uploaded_sheets else 'JUN 26'
                    except Exception:
                        default_sheet = 'JUN 26'
                        
                    new_wb.active_sheet = default_sheet
                    new_wb.save(update_fields=['active_sheet'])
                    
                    # Deactivate existing active workbook
                    BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
                    new_wb.is_active = True
                    new_wb.save(update_fields=['is_active'])
                    
                    messages.success(request, f"Uploaded workbook '{upload.name}' is now the active BTPL shipment workbook.")
                    return redirect('btpl_settings')
                else:
                    messages.error(request, "Please choose a valid .xlsx file to upload.")
    else:
        form = BtplWorkbookUploadForm(initial={'active_sheet': current_sheet_name}, sheets=sheets)
        
    context = {
        'active': 'btpl',
        'form': form,
        'wb': wb_obj,
        'current_sheet': current_sheet_name,
        'sheets': sheets,
    }
    return render(request, 'core/portal/btpl_settings.html', context)


@staff_required
@tool_permission_required('attendance')
@never_cache
def attendance_sheet(request):
    from django.urls import reverse
    wb_obj, file_path, sheet_name = attendance.get_active_attendance_workbook()
    
    if not file_path or not os.path.exists(file_path):
        return render(request, 'core/portal/attendance_form.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    if request.method == 'POST':
        try:
            modified = attendance.save_attendance(file_path, sheet_name, request.POST)
            if modified:
                ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_ATTENDANCE,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"Update: {sheet_name}",
                    detail=f"Updated attendance sheet '{sheet_name}'"
                )
                messages.success(request, "Attendance updated successfully.")
                return redirect(reverse('attendance_sheet') + '?success=1')
            else:
                messages.info(request, "No changes were made to the attendance sheet.")
                return redirect('attendance_sheet')
        except Exception as e:
            messages.error(request, f"Failed to save attendance: {e}")
            return redirect('attendance_sheet')
            
    # Load sheet data
    data = attendance.get_attendance_data(file_path, sheet_name)
    if not data:
        return render(request, 'core/portal/attendance_form.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    days_list = list(range(1, data['days_in_month'] + 1))
    days_info_list = []
    for d in days_list:
        days_info_list.append({
            'day': d,
            'is_holiday': data['day_headers'][d]['is_holiday']
        })
        
    employees_list = []
    for emp in data['employees']:
        days_attendance = []
        for d in days_list:
            days_attendance.append({
                'day': d,
                'status': emp['attendance'].get(d, ''),
                'is_holiday': data['day_headers'][d]['is_holiday']
            })
        employees_list.append({
            'row_idx': emp['row_idx'],
            'name': emp['name'],
            'phone': emp['phone'],
            'days_attendance': days_attendance
        })
        
    context = {
        'active': 'attendance',
        'sheet_name': data['sheet_name'],
        'days_in_month': data['days_in_month'],
        'days_info_list': days_info_list,
        'employees': employees_list,
        'sheet_names': data['sheet_names'],
        'wb': wb_obj,
        'success': request.GET.get('success') == '1',
    }
    return render(request, 'core/portal/attendance_form.html', context)


@staff_required
@tool_permission_required('attendance')
def attendance_download(request):
    wb_obj, file_path, sheet_name = attendance.get_active_attendance_workbook()
    if not file_path or not os.path.exists(file_path):
        raise Http404("Attendance workbook file not found.")
        
    filename = wb_obj.original_name if wb_obj else "Attendance_Sheet.xlsx"
    response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_required
@tool_permission_required('attendance')
def attendance_settings(request):
    """Manage active Attendance workbook and active sheet tab."""
    wb_obj, file_path, current_sheet_name = attendance.get_active_attendance_workbook()
    
    # Read sheets from Excel if file_path exists
    sheets = []
    if file_path and os.path.exists(file_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
        except Exception:
            sheets = ['JUNE 2026']
    else:
        sheets = ['JUNE 2026']
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'remove':
            AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
            if AttendanceWorkbook.objects.count() == 0:
                AttendanceWorkbook.objects.create(
                    original_name='Attendance_Sheet.xlsx',
                    active_sheet='JUNE 2026',
                    uploaded_by=request.user,
                    is_active=False
                )
            messages.success(request, "Attendance sheet removed entirely. Portal is now in empty state.")
            return redirect('attendance_settings')
            
        elif action == 'load_default':
            import shutil
            root_file_path = os.path.join(settings.BASE_DIR, 'Attendance_Sheet.xlsx')
            target_dir = os.path.join(settings.MEDIA_ROOT, 'attendance')
            os.makedirs(target_dir, exist_ok=True)
            dest_path = os.path.join(target_dir, 'Attendance_Sheet.xlsx')
            shutil.copy2(root_file_path, dest_path)
            
            AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
            
            wb_obj = AttendanceWorkbook.objects.create(
                original_name='Attendance_Sheet.xlsx',
                active_sheet='JUNE 2026',
                uploaded_by=request.user,
                is_active=True
            )
            wb_obj.file.name = 'attendance/Attendance_Sheet.xlsx'
            wb_obj.save(update_fields=['file'])
            
            try:
                wb = openpyxl.load_workbook(dest_path, read_only=True)
                if wb.sheetnames:
                    wb_obj.active_sheet = wb.sheetnames[-1]
                    wb_obj.save(update_fields=['active_sheet'])
            except Exception:
                pass
                
            messages.success(request, "Default attendance workbook loaded successfully.")
            return redirect('attendance_settings')
            
        elif action == 'change_sheet':
            form = AttendanceWorkbookUploadForm(request.POST, sheets=sheets)
            if form.is_valid():
                sheet_sel = form.cleaned_data.get('active_sheet')
                if wb_obj:
                    wb_obj.active_sheet = sheet_sel
                    wb_obj.save(update_fields=['active_sheet'])
                else:
                    import shutil
                    root_file_path = os.path.join(settings.BASE_DIR, 'Attendance_Sheet.xlsx')
                    target_dir = os.path.join(settings.MEDIA_ROOT, 'attendance')
                    os.makedirs(target_dir, exist_ok=True)
                    dest_path = os.path.join(target_dir, 'Attendance_Sheet.xlsx')
                    shutil.copy2(root_file_path, dest_path)
                    
                    wb_obj = AttendanceWorkbook.objects.create(
                        original_name='Attendance_Sheet.xlsx',
                        active_sheet=sheet_sel,
                        uploaded_by=request.user,
                        is_active=True
                    )
                    wb_obj.file.name = 'attendance/Attendance_Sheet.xlsx'
                    wb_obj.save(update_fields=['file'])
                    
                messages.success(request, f"Active sheet tab changed to '{sheet_sel}'.")
                return redirect('attendance_sheet')
                
        elif action == 'upload':
            form = AttendanceWorkbookUploadForm(request.POST, request.FILES, sheets=sheets)
            if form.is_valid():
                upload = request.FILES.get('workbook')
                if upload:
                    new_wb = AttendanceWorkbook.objects.create(
                        file=upload, original_name=upload.name,
                        uploaded_by=request.user, is_active=False
                    )
                    
                    try:
                        import openpyxl
                        temp_wb = openpyxl.load_workbook(new_wb.file.path, read_only=True)
                        uploaded_sheets = temp_wb.sheetnames
                        default_sheet = uploaded_sheets[-1] if uploaded_sheets else 'JUNE 2026'
                    except Exception:
                        default_sheet = 'JUNE 2026'
                        
                    new_wb.active_sheet = default_sheet
                    new_wb.save(update_fields=['active_sheet'])
                    
                    AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
                    new_wb.is_active = True
                    new_wb.save(update_fields=['is_active'])
                    
                    messages.success(request, f"Uploaded workbook '{upload.name}' is now the active attendance workbook.")
                    return redirect('attendance_settings')
                else:
                    messages.error(request, "Please choose a valid .xlsx file to upload.")
    else:
        form = AttendanceWorkbookUploadForm(initial={'active_sheet': current_sheet_name}, sheets=sheets)
        
    context = {
        'active': 'attendance',
        'form': form,
        'wb': wb_obj,
        'current_sheet': current_sheet_name,
        'sheets': sheets,
    }
    return render(request, 'core/portal/attendance_settings.html', context)