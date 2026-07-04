from django.core.management.base import BaseCommand
from core.models import Shipment
from core.models import BtplWorkbook
from core.btpl import get_column_mapping
import openpyxl
from django.utils import timezone
import dateutil.parser
import datetime

class Command(BaseCommand):
    help = 'Reconcile BTPL shipments between the active Excel workbook and the PostgreSQL database.'

    def handle(self, *args, **options):
        wb_obj = BtplWorkbook.active()
        if not wb_obj:
            self.stdout.write(self.style.ERROR("No active BTPL workbook found."))
            return

        file_path = wb_obj.file.path
        sheet_name = wb_obj.active_sheet

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"Sheet '{sheet_name}' not found in BTPL workbook."))
                return

            sheet = wb[sheet_name]
            mapping = get_column_mapping(sheet)
            
            if not mapping:
                self.stdout.write(self.style.ERROR("Failed to map columns for BTPL workbook."))
                return

            excel_records = {}
            # Read from row 2
            for row_idx in range(2, sheet.max_row + 1):
                def get_val(key):
                    col = mapping.get(key)
                    if col:
                        return sheet.cell(row=row_idx, column=col).value
                    return None

                pickup_date = get_val('pickup_date')
                lr_number = get_val('lr_number')
                name = get_val('name')

                if not pickup_date and not lr_number and not name:
                    continue
                
                if isinstance(pickup_date, str):
                    try:
                        pickup_date = dateutil.parser.parse(pickup_date).date()
                    except (ValueError, TypeError, dateutil.parser.ParserError):
                        pickup_date = None
                elif isinstance(pickup_date, datetime.datetime):
                    pickup_date = pickup_date.date()

                lr = str(lr_number or '')
                source_key = f"{lr}|{pickup_date.isoformat() if pickup_date else ''}"
                if source_key != "|":
                    excel_records[source_key] = row_idx
            
            wb.close()
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error reading Excel: {e}"))
            return

        db_records = set(Shipment.objects.filter(shipment_type='BTPL').values_list('source_key', flat=True))
        excel_keys = set(excel_records.keys())

        missing_in_db = excel_keys - db_records
        extra_in_db = db_records - excel_keys

        if not missing_in_db and not extra_in_db:
            self.stdout.write(self.style.SUCCESS("Reconciliation successful: 0 differences found between Excel and Database."))
        else:
            self.stdout.write(self.style.WARNING(f"Reconciliation found differences!"))
            if missing_in_db:
                self.stdout.write(self.style.ERROR(f"Missing in DB ({len(missing_in_db)} records):"))
                for key in list(missing_in_db)[:10]:
                    self.stdout.write(f"  - {key} (Excel Row {excel_records[key]})")
                if len(missing_in_db) > 10:
                    self.stdout.write("  ... (truncated)")
            if extra_in_db:
                self.stdout.write(self.style.ERROR(f"Extra in DB ({len(extra_in_db)} records):"))
                for key in list(extra_in_db)[:10]:
                    self.stdout.write(f"  - {key}")
                if len(extra_in_db) > 10:
                    self.stdout.write("  ... (truncated)")

