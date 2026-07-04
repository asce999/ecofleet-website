import os
import sys
import datetime
import dateutil.parser
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from core.models import Shipment, FtlWorkbook
from core.ftl import get_column_mapping


class Command(BaseCommand):
    help = 'Reconciles the active FTL workbook against the database and reports diffs.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--workbook-path',
            type=str,
            help='Path to the FTL workbook. Defaults to the active FtlWorkbook.',
            required=False,
        )

    def handle(self, *args, **options):
        workbook_path = options.get('workbook_path')
        sheet_name = 'Sheet1'

        if not workbook_path:
            wb_obj = FtlWorkbook.active()
            if wb_obj and wb_obj.file and os.path.exists(wb_obj.file.path):
                workbook_path = wb_obj.file.path
                sheet_name = wb_obj.active_sheet
            else:
                workbook_path = os.path.join(settings.BASE_DIR, 'efe_data', 'FTL_Shipment_Tracker.xlsx')

        if not os.path.exists(workbook_path):
            self.stderr.write(f"Workbook not found at {workbook_path}")
            sys.exit(1)

        self.stdout.write(f"Reconciling FTL against {workbook_path} (sheet {sheet_name})")

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            self.stderr.write(f"Sheet '{sheet_name}' not found in workbook.")
            sys.exit(1)

        sheet = wb[sheet_name]
        mapping = get_column_mapping(sheet)
        if not mapping:
            self.stderr.write("Failed to map columns for FTL workbook.")
            sys.exit(1)

        excel_rows = {}
        for row_idx in range(2, sheet.max_row + 1):
            def get_val(key):
                col = mapping.get(key)
                if col:
                    return sheet.cell(row=row_idx, column=col).value
                return None
            
            booking_date_raw = get_val('booking_date')
            lr_number_raw = get_val('lr_number')
            vehicle_number_raw = get_val('vehicle_number')

            if not booking_date_raw and not lr_number_raw and not vehicle_number_raw:
                continue

            booking_date = booking_date_raw
            if isinstance(booking_date, str):
                try:
                    booking_date = dateutil.parser.parse(booking_date).date()
                except (ValueError, TypeError, dateutil.parser.ParserError):
                    booking_date = None
            elif isinstance(booking_date, datetime.datetime):
                booking_date = booking_date.date()

            lr_number_str = str(lr_number_raw or '').strip().upper()
            source_key = f"{lr_number_str}|{booking_date.isoformat() if booking_date else ''}"
            
            if not lr_number_str and not booking_date:
                continue

            excel_rows[source_key] = {
                'row_idx': row_idx,
                'origin': str(get_val('origin') or '').strip(),
                'destination': str(get_val('destination') or '').strip(),
                'dispatch_date': booking_date,
                'lr_number': lr_number_str,
                'vehicle_number': str(vehicle_number_raw or '').strip().upper(),
            }

        wb.close()

        db_shipments = Shipment.objects.filter(shipment_type='FTL')
        db_rows = {}
        for s in db_shipments:
            db_rows[s.source_key] = {
                'origin': s.origin,
                'destination': s.destination,
                'dispatch_date': s.dispatch_date,
                'lr_number': s.metadata.get('lr_number', '').strip().upper(),
                'vehicle_number': s.vehicle.registration_number if s.vehicle else '',
            }

        excel_keys = set(excel_rows.keys())
        db_keys = set(db_rows.keys())

        matched_rows = 0
        mismatched_rows = 0
        mismatched_fields_count = 0
        only_in_sheet = excel_keys - db_keys
        only_in_db = db_keys - excel_keys
        common_keys = excel_keys & db_keys
        
        mismatch_details = []

        for key in common_keys:
            ex_row = excel_rows[key]
            db_row = db_rows[key]
            
            diffs = []
            if ex_row['origin'] != db_row['origin']:
                diffs.append(f"origin: '{ex_row['origin']}' != '{db_row['origin']}'")
            if ex_row['destination'] != db_row['destination']:
                diffs.append(f"destination: '{ex_row['destination']}' != '{db_row['destination']}'")
            if ex_row['dispatch_date'] != db_row['dispatch_date']:
                diffs.append(f"dispatch_date: '{ex_row['dispatch_date']}' != '{db_row['dispatch_date']}'")
            if ex_row['lr_number'] != db_row['lr_number']:
                diffs.append(f"lr_number: '{ex_row['lr_number']}' != '{db_row['lr_number']}'")
            if ex_row['vehicle_number'] != db_row['vehicle_number']:
                diffs.append(f"vehicle_number: '{ex_row['vehicle_number']}' != '{db_row['vehicle_number']}'")
                
            if diffs:
                mismatched_rows += 1
                mismatched_fields_count += len(diffs)
                mismatch_details.append(f"Key {key} (Row {ex_row['row_idx']}): " + ", ".join(diffs))
            else:
                matched_rows += 1

        self.stdout.write("\n=== Reconciliation Report ===")
        self.stdout.write(f"Total rows in Sheet: {len(excel_keys)}")
        self.stdout.write(f"Total rows in DB:    {len(db_keys)}")
        self.stdout.write(f"Matched rows:        {matched_rows}")
        self.stdout.write(f"Mismatched fields:   {mismatched_fields_count}")
        self.stdout.write(f"Rows only in Sheet:  {len(only_in_sheet)}")
        self.stdout.write(f"Rows only in DB:     {len(only_in_db)}")

        has_diffs = bool(mismatched_rows > 0 or only_in_sheet or only_in_db)

        if has_diffs:
            self.stdout.write("\n--- Details ---")
            for m in mismatch_details:
                self.stdout.write(f"Mismatch - {m}")
            for k in list(only_in_sheet)[:10]:
                self.stdout.write(f"Missing in DB - Key {k} (Row {excel_rows[k]['row_idx']})")
            if len(only_in_sheet) > 10:
                self.stdout.write(f"... and {len(only_in_sheet) - 10} more missing in DB")
                
            for k in list(only_in_db)[:10]:
                self.stdout.write(f"Missing in Sheet - Key {k}")
            if len(only_in_db) > 10:
                self.stdout.write(f"... and {len(only_in_db) - 10} more missing in Sheet")

            self.stdout.write("\nReconciliation FAILED.")
            sys.exit(1)
        else:
            self.stdout.write("\nReconciliation SUCCESSFUL.")
            sys.exit(0)
