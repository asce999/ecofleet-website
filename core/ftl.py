import datetime
import openpyxl
import os
from openpyxl.utils import get_column_letter, column_index_from_string

HEADER_MAP = {
    'booking_date': ['Date of Booking', 'Booking Date', 'Date'],
    'etd': ['ETD'],
    'delivery_date': ['Date of Delivery', 'Delivery Date', 'Actual Delivery Date'],
    'consignor': ['Consignor'],
    'origin': ['From Location', 'From', 'Origin'],
    'consignee': ['Consignee'],
    'lr_number': ['LR Number', 'LR No.', 'LR NO'],
    'destination': ['To Location', 'To', 'Destination'],
    'vehicle_number': ['Vehicle Number', 'Vehicle No.', 'Vehicle No', 'Vehicle'],
    'vendor': ['Vendor', 'Vendor Name'],
}

def get_column_mapping(sheet):
    """
    Scans the first row of the sheet to map standard keys to 1-based column indices.
    """
    mapping = {}
    for col in range(1, 40): # Scan up to 40 columns
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        val_str = str(val).strip().lower()
        # Find which key in HEADER_MAP matches
        for key, variations in HEADER_MAP.items():
            if any(var.strip().lower() == val_str for var in variations):
                mapping[key] = col
                break
                
    # Also find 'SR. NO.' or '#' for serial number column
    for col in range(1, 40):
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        val_str = str(val).strip().lower()
        if val_str in ['sr. no.', '#', 'serial', 'sr no', 'sr.no.', 's.no.']:
            mapping['serial'] = col
            break
            
    return mapping

def find_totals_row(sheet, mapping):
    """
    Finds the first row containing total summaries or sum formulas.
    """
    # FTL sheets typically don't have totals, but scan up to min(500, max_row+2) for total keywords
    for r in range(2, min(500, sheet.max_row + 2)):
        col1_val = sheet.cell(row=r, column=1).value
        if isinstance(col1_val, str) and any(x in col1_val.upper() for x in ['TOTAL', 'GRAND TOTAL', 'SUM']):
            return r
    return sheet.max_row + 1

def find_next_ftl_row(file_path, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    # Check for empty rows by verifying major identification and logistics fields
    for r in range(2, sheet.max_row + 2):
        has_data = False
        for key in ['booking_date', 'lr_number', 'consignee', 'vehicle_number']:
            col = mapping.get(key)
            if col:
                val = sheet.cell(row=r, column=col).value
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
        if not has_data:
            return r
    return sheet.max_row + 1

def derive_status(etd, delivery_date):
    if delivery_date is not None and str(delivery_date).strip() != "":
        return "Delivered"
    elif etd is not None and str(etd).strip() != "":
        return "In Transit"
    else:
        return "Booked"

def get_ftl_row_values(file_path, row, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    
    def get_val(key):
        col = mapping.get(key)
        if col is None:
            return None
        return sheet.cell(row=row, column=col).value
        
    return {
        'row_num': row,
        'booking_date': get_val('booking_date'),
        'etd': get_val('etd'),
        'delivery_date': get_val('delivery_date'),
        'consignor': get_val('consignor'),
        'origin': get_val('origin'),
        'consignee': get_val('consignee'),
        'lr_number': get_val('lr_number'),
        'destination': get_val('destination'),
        'vehicle_number': get_val('vehicle_number'),
        'vendor': get_val('vendor'),
    }

def add_ftl_shipment(file_path, row_data, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    row = row_data['row_num']
    
    mapping = get_column_mapping(sheet)
    
    def clean(val):
        if val == "" or val is None:
            return None
        if isinstance(val, str):
            val_strip = val.strip()
            if val_strip == "":
                return None
            return val_strip
        return val

    def write_val(key, val):
        col = mapping.get(key)
        if col is not None:
            sheet.cell(row=row, column=col).value = clean(val)

    write_val('booking_date', row_data['booking_date'])
    write_val('etd', row_data['etd'])
    write_val('delivery_date', row_data['delivery_date'])
    write_val('consignor', row_data['consignor'])
    write_val('origin', row_data['origin'])
    write_val('consignee', row_data['consignee'])
    write_val('lr_number', row_data['lr_number'])
    write_val('destination', row_data['destination'])
    write_val('vehicle_number', row_data['vehicle_number'])
    write_val('vendor', row_data['vendor'])
    
    # Auto-populate serial number if empty
    serial_col = mapping.get('serial')
    if serial_col is not None:
        current_serial = sheet.cell(row=row, column=serial_col).value
        if current_serial is None or str(current_serial).strip() == "":
            sheet.cell(row=row, column=serial_col).value = row - 1
            
    wb.save(file_path)

def evaluate_cell(sheet, row, col, memo=None):
    if memo is None:
        memo = {}
    
    cell_id = (row, col)
    if cell_id in memo:
        return memo[cell_id]
        
    val = sheet.cell(row=row, column=col).value
    # No formulas expected in FTL rows, but handle standard dates & values
    if isinstance(val, (int, float, datetime.datetime, datetime.date)):
        memo[cell_id] = val
        return val
    elif isinstance(val, str):
        if val.startswith('='):
            # Fallback evaluation for simple formulas if present
            memo[cell_id] = val
            return val
        try:
            if '.' in val:
                fval = float(val)
                memo[cell_id] = fval
                return fval
            else:
                ival = int(val)
                memo[cell_id] = ival
                return ival
        except ValueError:
            pass
    memo[cell_id] = val
    return val

def get_ftl_preview(file_path, sheet_name='Sheet1', page=1, page_size=20, sheet=None, mapping=None, memo=None):
    """Return paginated preview data."""
    if sheet is None:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            return {'columns': [], 'rows': [], 'total_rows': 0, 'page': 1, 'page_size': page_size, 'total_pages': 0}
        sheet = wb[sheet_name]

    if mapping is None:
        mapping = get_column_mapping(sheet)
    if memo is None:
        memo = {}

    max_col = max(mapping.values()) if mapping else 11

    def format_cell(val):
        if val is None:
            return ""
        elif isinstance(val, (datetime.datetime, datetime.date)):
            return val.strftime('%d-%b-%y')
        elif isinstance(val, float):
            return str(int(val)) if val.is_integer() else f"{val:.2f}"
        else:
            return str(val)

    # Get header row
    columns = []
    for c in range(1, max_col + 1):
        val = evaluate_cell(sheet, 1, c, memo)
        columns.append(format_cell(val))

    # Collect all non-empty data rows (row 2 onwards)
    all_data_rows = []
    for r in range(2, sheet.max_row + 1):
        row_vals = []
        has_data = False
        for c in range(1, max_col + 1):
            val = evaluate_cell(sheet, r, c, memo)
            formatted = format_cell(val)
            row_vals.append(formatted)
            # Check if any cell has non-empty values besides the serial number
            if formatted != "" and c != mapping.get('serial'):
                has_data = True
        if has_data:
            all_data_rows.append({'row_num': r, 'cells': row_vals})

    total_rows = len(all_data_rows)
    total_pages = max(1, (total_rows + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))

    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_rows = all_data_rows[start_idx:end_idx]

    return {
        'columns': columns,
        'rows': page_rows,
        'total_rows': total_rows,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }

def get_ftl_page_data(file_path, sheet_name='Sheet1', target_row=None, page=1, page_size=20):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return None

    sheet = wb[sheet_name]
    mapping = get_column_mapping(sheet)
    totals_row = find_totals_row(sheet, mapping)
    memo = {}

    # Find next empty row
    auto_next_row = None
    for r in range(2, sheet.max_row + 2):
        has_data = False
        for key in ['booking_date', 'lr_number', 'consignee', 'vehicle_number']:
            col = mapping.get(key)
            if col:
                val = sheet.cell(row=r, column=col).value
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
        if not has_data:
            auto_next_row = r
            break
    if not auto_next_row:
        auto_next_row = sheet.max_row + 1

    next_row = None
    if target_row and 2 <= target_row < totals_row:
        next_row = target_row
    else:
        next_row = auto_next_row

    row_values = {}
    if next_row:
        def get_val(key):
            col = mapping.get(key)
            if col is None:
                return None
            return sheet.cell(row=next_row, column=col).value

        row_values = {
            'row_num': next_row,
            'booking_date': get_val('booking_date'),
            'etd': get_val('etd'),
            'delivery_date': get_val('delivery_date'),
            'consignor': get_val('consignor'),
            'origin': get_val('origin'),
            'consignee': get_val('consignee'),
            'lr_number': get_val('lr_number'),
            'destination': get_val('destination'),
            'vehicle_number': get_val('vehicle_number'),
            'vendor': get_val('vendor'),
        }

    preview = get_ftl_preview(
        file_path, sheet_name=sheet_name,
        page=page, page_size=page_size,
        sheet=sheet, mapping=mapping, memo=memo
    )

    return {
        'mapping': mapping,
        'totals_row': totals_row,
        'next_row': next_row,
        'row_values': row_values,
        'preview': preview,
    }

def clear_ftl_row(file_path, row, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    # Clear all cells except serial number
    for key, col in mapping.items():
        if key != 'serial':
            sheet.cell(row=row, column=col).value = None
            
    wb.save(file_path)

def get_cached_ftl_metrics(ftl_wb_obj, ftl_file_path, ftl_sheet_name):
    import os
    from django.core.cache import cache
    from pathlib import Path
    
    ftl_total = 0
    ftl_delivered = 0
    ftl_in_transit = 0
    ftl_vendors = 0
    
    cache_key = None
    if ftl_wb_obj and ftl_file_path and Path(ftl_file_path).exists():
        mtime = os.path.getmtime(ftl_file_path)
        cache_key = f'ftl_metrics_active_{ftl_wb_obj.id}_{mtime}'
        
    cached_metrics = cache.get(cache_key) if cache_key else None
    if cached_metrics:
        return cached_metrics
        
    if ftl_file_path and Path(ftl_file_path).exists():
        try:
            import openpyxl
            wb_ftl = openpyxl.load_workbook(ftl_file_path, read_only=True)
            if ftl_sheet_name in wb_ftl.sheetnames:
                sheet_ftl = wb_ftl[ftl_sheet_name]
                mapping_ftl = get_column_mapping(sheet_ftl)
                vendors_set = set()
                del_col = mapping_ftl.get('delivery_date')
                etd_col = mapping_ftl.get('etd')
                vendor_col = mapping_ftl.get('vendor')
                
                for r in range(2, sheet_ftl.max_row + 1):
                    has_data = False
                    for key in ['booking_date', 'lr_number', 'consignee', 'vehicle_number']:
                        col = mapping_ftl.get(key)
                        if col:
                            val = sheet_ftl.cell(row=r, column=col).value
                            if val is not None and str(val).strip() != '':
                                has_data = True
                                break
                    if not has_data:
                        continue
                        
                    ftl_total += 1
                    del_val = sheet_ftl.cell(row=r, column=del_col).value if del_col else None
                    etd_val = sheet_ftl.cell(row=r, column=etd_col).value if etd_col else None
                    
                    status = derive_status(etd_val, del_val)
                    if status == 'Delivered':
                        ftl_delivered += 1
                    elif status == 'In Transit':
                        ftl_in_transit += 1
                        
                    v_val = sheet_ftl.cell(row=r, column=vendor_col).value if vendor_col else None
                    if v_val is not None and str(v_val).strip() != '':
                        vendors_set.add(str(v_val).strip())
                ftl_vendors = len(vendors_set)
                
                if cache_key:
                    cache.set(cache_key, (ftl_total, ftl_delivered, ftl_in_transit, ftl_vendors), timeout=900)
        except Exception:
            pass
            
    return ftl_total, ftl_delivered, ftl_in_transit, ftl_vendors

