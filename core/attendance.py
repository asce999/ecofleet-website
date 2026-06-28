import calendar
import datetime
import openpyxl
import os
from django.conf import settings
from openpyxl.utils import get_column_letter

PAID_CODES = {'P', 'CL', 'SL', 'PL', 'WO', 'H'}
HALF_DAY_CODES = {'HD'}
ABSENT_CODES = {'A'}

def get_month_year_from_sheet(sheet_name):
    parts = sheet_name.strip().split()
    if len(parts) == 2:
        month_name, year_str = parts
        try:
            year = int(year_str)
            try:
                month = datetime.datetime.strptime(month_name, "%B").month
            except ValueError:
                month = datetime.datetime.strptime(month_name, "%b").month
            return year, month
        except Exception:
            pass
    raise ValueError(f"Could not extract a valid month and year from sheet name: '{sheet_name}'")

def get_days_in_sheet(year, month):
    return calendar.monthrange(year, month)[1]

def get_active_attendance_workbook():
    from core.models import AttendanceWorkbook
    wb_count = AttendanceWorkbook.objects.count()
    if wb_count > 0:
        wb_obj = AttendanceWorkbook.active()
        if not wb_obj:
            return None, None, None
        return wb_obj, wb_obj.file.path, wb_obj.active_sheet
    else:
        file_path = os.path.join(settings.BASE_DIR, 'Attendance_Sheet.xlsx')
        sheet_name = 'JUNE 2026'
        if os.path.exists(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                if wb.sheetnames:
                    sheet_name = wb.sheetnames[-1]
            except Exception:
                pass
        return None, file_path, sheet_name

def get_attendance_data(file_path, sheet_name, manual_year=None, manual_month=None):
    if not os.path.exists(file_path):
        return None
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[-1] if wb.sheetnames else None
        if not sheet_name:
            return None
            
    try:
        if manual_year and manual_month:
            year, month = int(manual_year), int(manual_month)
        else:
            year, month = get_month_year_from_sheet(sheet_name)
    except ValueError:
        return {
            'requires_manual_date': True,
            'sheet_name': sheet_name,
            'sheet_names': wb.sheetnames,
        }

    sheet = wb[sheet_name]
    
    # 1. Determine number of days in the month
    days_in_month = get_days_in_sheet(year, month)
    
    # 2. Find STAFF NAME and PHONE NUMBER column indices
    staff_col = 1
    phone_col = None
    dept_col = None
    extra_days_col = None
    payable_days_col = None
    emp_id_col = None
    pf_no_col = None
    esic_no_col = None
    uan_no_col = None
    
    for c in range(1, sheet.max_column + 1):
        h_val = sheet.cell(row=1, column=c).value
        if h_val:
            h_str = str(h_val).strip().lower()
            if h_str == 'staff name' or h_str == 'name':
                staff_col = c
            elif h_str == 'phone number' or h_str == 'phone':
                phone_col = c
            elif h_str == 'department' or h_str == 'dept':
                dept_col = c
            elif h_str == 'payable days' or h_str == 'payable_days':
                payable_days_col = c
            elif h_str == 'extra days' or h_str == 'overtime':
                extra_days_col = c
            elif h_str == 'empid' or h_str == 'emp id' or h_str == 'employee id':
                emp_id_col = c
            elif h_str == 'pf no' or h_str == 'pf number':
                pf_no_col = c
            elif h_str == 'esic no' or h_str == 'esic number':
                esic_no_col = c
            elif h_str == 'uan no' or h_str == 'uan number':
                uan_no_col = c

    # 3. Identify holidays/weekly offs for days
    day_headers = {}
    for d in range(1, days_in_month + 1):
        col_idx = d + staff_col
        h_val = sheet.cell(row=1, column=col_idx).value
        is_holiday = False
        if h_val:
            if str(h_val).strip().upper() == 'HD':
                is_holiday = True
        day_headers[d] = {
            'col_idx': col_idx,
            'is_holiday': is_holiday,
            'header_val': h_val if h_val is not None else ""
        }

    # 4. Extract employees and their attendance
    employees = []
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(row=r, column=staff_col).value
        if name is None or str(name).strip() == "":
            continue
            
        name = str(name).strip()
        phone = ""
        department = ""
        payable_days_from_sheet = None
        extra_days = 0.0
        emp_id = ""
        pf_no = ""
        esic_no = ""
        uan_no = ""
        
        if phone_col:
            phone_val = sheet.cell(row=r, column=phone_col).value
            if phone_val is not None:
                phone = str(phone_val).strip()
                
        if dept_col:
            val = sheet.cell(row=r, column=dept_col).value
            if val is not None: department = str(val).strip()
            
        if payable_days_col:
            val = sheet.cell(row=r, column=payable_days_col).value
            if val is not None:
                try: payable_days_from_sheet = float(val)
                except ValueError: pass
                
        if extra_days_col:
            val = sheet.cell(row=r, column=extra_days_col).value
            if val is not None:
                try: extra_days = float(val)
                except ValueError: extra_days = 0.0
                
        if emp_id_col:
            val = sheet.cell(row=r, column=emp_id_col).value
            if val is not None: emp_id = str(val).strip()
            
        if pf_no_col:
            val = sheet.cell(row=r, column=pf_no_col).value
            if val is not None: pf_no = str(val).strip()
            
        if esic_no_col:
            val = sheet.cell(row=r, column=esic_no_col).value
            if val is not None: esic_no = str(val).strip()
            
        if uan_no_col:
            val = sheet.cell(row=r, column=uan_no_col).value
            if val is not None: uan_no = str(val).strip()
                
        attendance = {}
        for d in range(1, days_in_month + 1):
            col_idx = day_headers[d]['col_idx']
            val = sheet.cell(row=r, column=col_idx).value
            status = ""
            if val is not None:
                status = str(val).strip().upper()
            attendance[d] = status
            
        employees.append({
            'row_idx': r,
            'name': name,
            'phone': phone,
            'department': department,
            'payable_days': payable_days_from_sheet,
            'extra_days': extra_days,
            'emp_id': emp_id,
            'pf_no': pf_no,
            'esic_no': esic_no,
            'uan_no': uan_no,
            'attendance': attendance
        })
        
    return {
        'sheet_name': sheet_name,
        'year': year,
        'month': month,
        'days_in_month': days_in_month,
        'day_headers': day_headers,
        'employees': employees,
        'sheet_names': wb.sheetnames
    }

def save_attendance(file_path, sheet_name, post_data):
    if not os.path.exists(file_path):
        raise FileNotFoundError("Attendance sheet file not found.")
        
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet tab '{sheet_name}' not found in workbook.")
        
    sheet = wb[sheet_name]
    
    # Get column mapping
    staff_col = 1
    for c in range(1, sheet.max_column + 1):
        h_val = sheet.cell(row=1, column=c).value
        if h_val and str(h_val).strip().lower() == 'staff name':
            staff_col = c
            break

    # Get valid bounds for server-side validation
    attendance_data = get_attendance_data(file_path, sheet_name)
    if not attendance_data:
        return False
    valid_rows = {emp['row_idx'] for emp in attendance_data.get('employees', [])}
    days_in_month = calendar.monthrange(attendance_data['year'], attendance_data['month'])[1]
    valid_days = set(range(1, days_in_month + 1))

    # Look for inputs like: attendance_{row_idx}_{day}
    modified = False
    for key, val in post_data.items():
        if key.startswith('attendance_'):
            parts = key.split('_')
            if len(parts) == 3:
                try:
                    row_idx = int(parts[1])
                    day = int(parts[2])
                    
                    if row_idx not in valid_rows or day not in valid_days:
                        continue  # silently skip malformed/out-of-bounds keys
                        
                    col_idx = day + staff_col
                    
                    # Clean the status value
                    status = str(val).strip().upper()
                    if status in ['', 'EMPTY', 'NONE']:
                        new_val = None
                    else:
                        new_val = status
                        
                    current_val = sheet.cell(row=row_idx, column=col_idx).value
                    # Check if actually modified to reduce writes
                    current_str = str(current_val).strip().upper() if current_val is not None else ""
                    new_str = new_val if new_val is not None else ""
                    if current_str != new_str:
                        sheet.cell(row=row_idx, column=col_idx).value = new_val
                        modified = True
                except (ValueError, TypeError):
                    pass
                    
    if modified:
        wb.save(file_path)
    return modified


def get_working_days(year, month):
    days_in_month = get_days_in_sheet(year, month)
    
    sundays = 0
    for day in range(1, days_in_month + 1):
        # weekday() returns 6 for Sunday
        if datetime.date(year, month, day).weekday() == 6:
            sundays += 1
            
    working_days = days_in_month - sundays
    return working_days, sundays

def calculate_salary_data(attendance_data):
    from core.models import EmployeeSalaryOverride, SalaryConfig
    from decimal import Decimal, ROUND_HALF_UP

    year = attendance_data['year']
    month = attendance_data['month']
    working_days, sundays = get_working_days(year, month)
    
    config = SalaryConfig.get_solo()
    
    employee_names = [emp['name'] for emp in attendance_data['employees']]
    overrides = EmployeeSalaryOverride.objects.filter(employee_name__in=employee_names)
    override_dict = {o.employee_name: o for o in overrides}
    
    results = []
    
    total_man_days = Decimal('0.00')
    total_net_payment = Decimal('0.00')
    total_cash_payment = Decimal('0.00')
    sum_total_a = Decimal('0.00')
    sum_total_b = Decimal('0.00')
    total_pt_tax = Decimal('0.00')
    
    eligible_depts = [d.strip() for d in config.other_allowance_eligible_departments.split(',') if d.strip()]
    
    def my_round(val):
        return val.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
    for emp in attendance_data['employees']:
        name = emp['name']
        dept = emp.get('department', '')
        extra_days_from_sheet = Decimal(str(emp.get('extra_days', 0.0)))
        
        override = override_dict.get(name)
        if override:
            adhoc_salary_increase_pct = override.adhoc_salary_increase_pct / Decimal('100.0')
            adhoc_allowance_monthly_amount = override.adhoc_allowance_monthly_amount
            advance = override.advance
            lwf = override.lwf
            other_deduction = override.other_deduction
            cash_payment = override.cash_payment
        else:
            adhoc_salary_increase_pct = Decimal('0.00')
            adhoc_allowance_monthly_amount = Decimal('0.00')
            advance = Decimal('0.00')
            lwf = Decimal('0.00')
            other_deduction = Decimal('0.00')
            cash_payment = Decimal('0.00')
            
        present_count = 0
        a_count = 0
        hd_count = 0
        
        for d, status in emp['attendance'].items():
            if status in PAID_CODES:
                present_count += 1
            elif status in ABSENT_CODES:
                a_count += 1
            elif status in HALF_DAY_CODES:
                hd_count += 1
                
        computed_payable_days = Decimal(present_count) + (Decimal(hd_count) * Decimal('0.5'))
        sheet_payable_days = emp.get('payable_days')
        if sheet_payable_days is not None:
            payable_days = Decimal(str(sheet_payable_days))
        else:
            payable_days = computed_payable_days
            
        extra_days = extra_days_from_sheet
        
        # 1. basic
        basic = my_round(payable_days * config.basic_rate_per_day)
        
        # 2. adhocSalaryIncrease
        adhoc_salary_increase = basic * adhoc_salary_increase_pct
        
        # 3. spAllowance
        sp_allowance = my_round(payable_days * config.sp_allowance_rate_per_day)
        
        # 4. adhocAllowance
        adhoc_allowance = my_round((adhoc_allowance_monthly_amount / Decimal(config.standard_month_days)) * payable_days)
        
        # 5. totalA
        total_a = basic + adhoc_salary_increase + sp_allowance + adhoc_allowance
        
        # 6. otherAllowance
        if dept in eligible_depts:
            other_allowance = my_round((basic + sp_allowance) * (config.other_allowance_pct / Decimal('100.0')) / Decimal(config.standard_month_days) * payable_days)
        else:
            other_allowance = Decimal('0.00')
            
        # 7. hra
        hra = my_round((basic + sp_allowance) * (config.hra_pct / Decimal('100.0')))
        
        # 8. leavePayment
        if payable_days > config.leave_payment_threshold_days:
            leave_payment = config.leave_payment_amount
        else:
            leave_payment = Decimal('0.00')
            
        # 9. extraPayment
        extra_payment = my_round(extra_days * config.extra_day_rate)
        
        # 10. totalB
        total_b = total_a + other_allowance + hra + leave_payment + extra_payment
        
        # 11. pf
        if total_a >= config.pf_wage_ceiling:
            pf = my_round(config.pf_wage_ceiling * (config.pf_rate / Decimal('100.0')))
        else:
            pf = my_round(total_a * (config.pf_rate / Decimal('100.0')))
            
        # 12. esicEmployee
        esic_employee = my_round(total_b * (config.esic_employee_rate / Decimal('100.0')))
        
        # 13. professionalTax
        if total_b <= config.pt_slab_1_max:
            pt = Decimal('0.00')
        elif total_b <= config.pt_slab_2_max:
            pt = Decimal('175.00')
        elif total_b > Decimal('10000.00'):
            pt = config.pt_slab_3_amount
        else:
            pt = Decimal('0.00')
            
        # 14. canteen
        canteen = my_round(payable_days * config.canteen_rate_per_day)
        
        # 15. subTotalDeductions
        sub_total_deductions = pf + esic_employee + pt + advance + lwf + canteen + other_deduction
        
        # 16. netPayment
        net_payment = my_round(total_b - sub_total_deductions)
        
        total_man_days += (payable_days + extra_days)
        sum_total_a += total_a
        sum_total_b += total_b
        total_pt_tax += pt
        total_net_payment += net_payment
        total_cash_payment += cash_payment
        
        results.append({
            'name': name,
            'department': dept,
            'emp_id': emp.get('emp_id', ''),
            'pf_no': emp.get('pf_no', ''),
            'esic_no': emp.get('esic_no', ''),
            'uan_no': emp.get('uan_no', ''),
            'payable_days': payable_days,
            'extra_days': extra_days,
            'basic': basic,
            'adhoc_salary_increase': adhoc_salary_increase,
            'sp_allowance': sp_allowance,
            'adhoc_allowance': adhoc_allowance,
            'total_a': total_a,
            'other_allowance': other_allowance,
            'hra': hra,
            'leave_payment': leave_payment,
            'extra_payment': extra_payment,
            'total_b': total_b,
            'pf': pf,
            'esic_employee': esic_employee,
            'pt': pt,
            'advance': advance,
            'lwf': lwf,
            'canteen': canteen,
            'other_deduction': other_deduction,
            'sub_total_deductions': sub_total_deductions,
            'net_payment': net_payment,
            'adhoc_salary_increase_pct': override.adhoc_salary_increase_pct if override else Decimal('0.00'),
            'adhoc_allowance_monthly_amount': override.adhoc_allowance_monthly_amount if override else Decimal('0.00'),
            'cash_payment': cash_payment
        })
        
    total_payment = total_net_payment + total_cash_payment
    total_esic_cost = my_round(sum_total_b * (config.reporting_esic_cost_rate / Decimal('100.0')))
    total_pf_cost = my_round(sum_total_a * (config.reporting_pf_cost_rate / Decimal('100.0')))
    total_cost = total_pt_tax + total_payment + total_esic_cost + total_pf_cost
    
    return {
        'working_days': working_days,
        'sundays': sundays,
        'total_employees': len(results),
        'total_man_days': total_man_days,
        'total_net_payment': total_net_payment,
        'total_payment': total_payment,
        'total_esic_cost': total_esic_cost,
        'total_pf_cost': total_pf_cost,
        'total_pt_tax': total_pt_tax,
        'total_cost': total_cost,
        'employees': results
    }

def get_sheet_names(file_path):
    from pathlib import Path
    if file_path and Path(file_path).exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            return wb.sheetnames
        except Exception:
            pass
    return []

