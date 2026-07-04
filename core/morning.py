"""
Morning Report — upload model (Delhivery only; DTDC removed).

Inputs (uploaded):
  • The Delhivery CSV export(s) emailed daily (one file per account, e.g. ecofleet4/ecofleet6).
    Each row's vehicle type is read from "Client Location/warehouse" (… 2 wheeler → 2W, else CV).
  • Yesterday's 2W master (.xlsx)  — sheets "2W Report" + "Pin-Code"
  • Yesterday's CV master (.xlsx)  — sheets "CV Report" + "Pin-Codes"

Output: updated 2W and CV master workbooks (only current-month rows kept).
"""
import io
import datetime
from datetime import timedelta
import warnings

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore")


class ReportError(Exception):
    """A user-facing problem with the uploaded files."""


MASTER_COLUMNS = [
    "S.NO",
    "Date of Shipping",
    "Booking Origin",
    "Booking Code",
    "Account of Customer - Consignor Name",
    "Lr No. ",
    "Mode",
    "Invoice Number",
    "Invoice Basic Value",
    "Invoice Net Value",
    "No. of Packages",
    "Consignee Name",
    "Consignee - City",
    "Consignee - Pincode",
    "Consignee - State",
    "Region",
    "Delivery Mode (Surface/ Air / Express)",
    "Charge Weight  in Kg",
    "Agreed TAT ",
    "Estimated Date of Delivery ( Based upon Agreed TAT)",
    "Date",
    "Days",
    "Status",
    "Actual Delivery Date",
    "Undelivered  ",
    "Remarks",
]


# ── Helpers ──────────────────────────────────────────────────────────
def _today():
    return pd.Timestamp(datetime.datetime.today().date())


def _report_date():
    return datetime.datetime.today().date() - timedelta(days=1)


def normalize_lr(series):
    """Strip whitespace and remove a trailing .0 from LR numbers."""
    return series.astype(str).str.strip().str.replace(r'\.0$', '', regex=True)


def find_pincode_col(df, candidates, label):
    stripped = {c.strip(): c for c in df.columns}
    for name in candidates:
        if name.strip() in stripped:
            return stripped[name.strip()]
    raise ReportError(f"Could not find the {label} column in the pincode sheet.")


def find_lr_col(df):
    for col in df.columns:
        normalised = str(col).strip().lower().replace(" ", "").replace(".", "")
        if normalised == "lrno":
            return col
    raise ReportError("Could not find the 'Lr No.' column in a master report sheet.")


def load_pincode_map(all_sheets, vehicle_type):
    sheet_name = "Pin-Code" if vehicle_type == "2W" else "Pin-Codes"
    if sheet_name not in all_sheets:
        raise ReportError(
            f'The {vehicle_type} master has no "{sheet_name}" sheet — is it the right file?')
    df = all_sheets[sheet_name].copy()
    df = df.astype(str)
    df.columns = df.columns.str.strip()

    pincode_col = find_pincode_col(df, ["Consignee - Pincode", "Pincode", "Pin Code", "PIN"], "Pincode")
    tat_col     = find_pincode_col(df, ["TAT", "Agreed TAT", "Agreed TAT ", "TAT (Days)"], "TAT")
    region_col  = find_pincode_col(df, ["Region"], "Region")
    state_col   = find_pincode_col(df, ["Consignee - State", "State", "Destination State"], "State")

    df[pincode_col] = df[pincode_col].str.strip()
    df = df.rename(columns={
        pincode_col: "Consignee - Pincode",
        tat_col:     "TAT",
        region_col:  "Region",
        state_col:   "Consignee - State",
    })
    df["TAT"] = (
        df["TAT"].astype(str).str.strip()
        .str.extract(r'(\d+\.?\d*)', expand=False)
        .pipe(pd.to_numeric, errors="coerce")
        .astype("Int64")
    )
    df = df.drop_duplicates(subset="Consignee - Pincode", keep="first")
    return df.set_index("Consignee - Pincode")[["TAT", "Region", "Consignee - State"]]


def clean_delhivery(file):
    """Clean one Delhivery CSV → normalised rows tagged with vehicle_type."""
    try:
        df = pd.read_csv(file, dtype={"LRN": str, "Pin code": str})
    except Exception as e:
        raise ReportError(f"Couldn't read a Delhivery CSV ({e}).")
    df.columns = df.columns.str.strip()
    if "LRN" not in df.columns or "Client Location/warehouse" not in df.columns:
        raise ReportError(
            "A Delhivery CSV is missing expected columns (LRN / Client Location/warehouse).")

    df["LRN"] = normalize_lr(df["LRN"])
    df["Pin code"] = df["Pin code"].astype(str).str.strip()

    df["No of boxes"] = pd.to_numeric(df["No of boxes"], errors="coerce").fillna(0).astype(int) - 1
    df["No of boxes"] = df["No of boxes"].clip(lower=0)

    df["Consignee name"] = df["Consignee name"].astype(str).str.split(",").str[0].str.strip()

    df["Status_mapped"] = df["Current Status"].apply(
        lambda x: "Delivered" if str(x).strip().lower() == "delivered" else "In Transit")

    df["Expected Date"]  = pd.to_datetime(df["Expected Date"],  errors="coerce").dt.normalize()
    df["Delivered Date"] = pd.to_datetime(df["Delivered Date"], errors="coerce").dt.normalize()
    df["Pickup Date"]    = pd.to_datetime(df["Pickup Date"],    errors="coerce").dt.normalize()

    # Use the logical report date (yesterday) to determine the target month
    report_date = _report_date()
    df = df[(df["Pickup Date"].dt.year == report_date.year) & (df["Pickup Date"].dt.month == report_date.month)]

    df["vehicle_type"] = df["Client Location/warehouse"].astype(str).apply(
        lambda x: "2W" if "2 wheeler" in x.lower() else "CV")

    df["_status_rank"] = df["Status_mapped"].apply(lambda x: 0 if x == "Delivered" else 1)
    df = (df.sort_values("_status_rank")
            .drop_duplicates(subset="LRN", keep="first")
            .drop(columns="_status_rank"))
    return df


def load_master(all_sheets, vehicle_type):
    sheet = "2W Report" if vehicle_type == "2W" else "CV Report"
    if sheet not in all_sheets:
        raise ReportError(
            f'The {vehicle_type} master has no "{sheet}" sheet — is it the right file?')
    df = all_sheets[sheet].copy()

    lr_col_actual = find_lr_col(df)
    if lr_col_actual != "Lr No. ":
        df = df.rename(columns={lr_col_actual: "Lr No. "})
    df["Lr No. "] = normalize_lr(df["Lr No. "])

    if "Consignor Name" in df.columns and "Account of Customer - Consignor Name" not in df.columns:
        df = df.rename(columns={"Consignor Name": "Account of Customer - Consignor Name"})
    if "ETD" in df.columns and "Estimated Date of Delivery ( Based upon Agreed TAT)" not in df.columns:
        df = df.rename(columns={"ETD": "Estimated Date of Delivery ( Based upon Agreed TAT)"})

    for _dc in ["Date of Shipping",
                "Estimated Date of Delivery ( Based upon Agreed TAT)",
                "Date", "Actual Delivery Date"]:
        if _dc in df.columns:
            df[_dc] = pd.to_datetime(df[_dc], errors="coerce").dt.normalize()

    date_col = "Date of Shipping"
    if date_col in df.columns:
        report_date = _report_date()
        df = df[(df[date_col].dt.year == report_date.year) & (df[date_col].dt.month == report_date.month)]

    df = df.drop_duplicates(subset="Lr No. ", keep="first")
    for col in MASTER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[MASTER_COLUMNS]


# ── Core update logic ────────────────────────────────────────────────
def update_existing_rows(master_df, report_df, lr_col_report, delivered_date_col):
    today = _today()
    etd_col = "Estimated Date of Delivery ( Based upon Agreed TAT)"
    report_lookup = report_df.set_index(lr_col_report)

    updated = 0
    for idx, row in master_df.iterrows():
        lr = str(row["Lr No. "]).strip()
        if lr not in report_lookup.index:
            continue
        rep = report_lookup.loc[lr]
        if isinstance(rep, pd.DataFrame):
            rep = rep.iloc[0]

        status = rep["Status_mapped"]
        master_df.at[idx, "Status"] = status

        if status == "Delivered":
            delivery_date = rep.get(delivered_date_col, pd.NaT)
            master_df.at[idx, "Actual Delivery Date"] = delivery_date if pd.notna(delivery_date) else pd.NaT
            master_df.at[idx, "Date"] = pd.NaT
            master_df.at[idx, "Days"] = None
        else:
            master_df.at[idx, "Date"] = today
            master_df.at[idx, "Actual Delivery Date"] = pd.NaT
            etd = row[etd_col]
            if pd.notna(etd):
                master_df.at[idx, "Days"] = (pd.Timestamp(etd) - today).days
        updated += 1
    return master_df, updated


def build_new_rows(report_df, pincode_map, vehicle_type, existing_max_sno):
    today = _today()
    consignor = "Piaggio (2Wheeler)" if vehicle_type == "2W" else "Piaggio (3Wheeler)"
    new_rows = []
    sno = existing_max_sno

    for _, rep in report_df.iterrows():
        sno += 1
        pincode = str(rep.get("Pin code", "")).strip()
        if pincode in pincode_map.index:
            tat_val = pincode_map.loc[pincode, "TAT"]
            tat = int(tat_val) if pd.notna(tat_val) else None
            region = pincode_map.loc[pincode, "Region"]
            state = pincode_map.loc[pincode, "Consignee - State"]
        else:
            tat = None
            state = rep.get("State", "")
            region = ""

        ship_date = rep["Pickup Date"]
        etd = (ship_date + timedelta(days=tat)) if (pd.notna(ship_date) and tat) else pd.NaT
        status = rep["Status_mapped"]
        raw_delivery = rep.get("Delivered Date", pd.NaT)
        actual_delivery = raw_delivery if (status == "Delivered" and pd.notna(raw_delivery)) else pd.NaT
        date_col_val = today if status == "In Transit" else pd.NaT
        days_val = (pd.Timestamp(etd) - today).days if (status == "In Transit" and pd.notna(etd)) else None

        new_rows.append({
            "S.NO": sno,
            "Date of Shipping": ship_date,
            "Booking Origin": "Baramati",
            "Booking Code": vehicle_type,
            "Account of Customer - Consignor Name": consignor,
            "Lr No. ": rep["LRN"],
            "Mode": "Delhivery",
            "Invoice Number": rep.get("Invoice Number", ""),
            "Invoice Basic Value": rep.get("Package Amount", None),
            "Invoice Net Value": rep.get("Package Amount", None),
            "No. of Packages": rep["No of boxes"],
            "Consignee Name": rep["Consignee name"],
            "Consignee - City": rep.get("Destination City", ""),
            "Consignee - Pincode": pincode,
            "Consignee - State": state,
            "Region": region,
            "Delivery Mode (Surface/ Air / Express)": "Surface",
            "Charge Weight  in Kg": rep.get("Weight", None),
            "Agreed TAT ": tat,
            "Estimated Date of Delivery ( Based upon Agreed TAT)": etd,
            "Date": date_col_val,
            "Days": days_val,
            "Status": status,
            "Actual Delivery Date": actual_delivery,
            "Undelivered  ": None,
            "Remarks": None,
        })
    return new_rows


# ── Save with formatting ─────────────────────────────────────────────
def save_master(master_df, pincode_df, vehicle_type):
    report_sheet  = "2W Report" if vehicle_type == "2W" else "CV Report"
    pincode_sheet = "Pin-Code"  if vehicle_type == "2W" else "Pin-Codes"

    master_df = master_df.drop_duplicates(subset="Lr No. ", keep="first").reset_index(drop=True)
    master_df["S.NO"] = range(1, len(master_df) + 1)
    for col in MASTER_COLUMNS:
        if col not in master_df.columns:
            master_df[col] = None
    master_df = master_df[MASTER_COLUMNS]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name=report_sheet, index=False)
        pincode_df.to_excel(writer, sheet_name=pincode_sheet, index=False)

        ws = writer.sheets[report_sheet]
        header_fill  = PatternFill("solid", fgColor="1F4E79")
        header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

        date_col_names = [
            "Date of Shipping",
            "Estimated Date of Delivery ( Based upon Agreed TAT)",
            "Date", "Actual Delivery Date",
        ]
        date_col_indices = [i for i, cell in enumerate(ws[1], 1) if cell.value in date_col_names]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in date_col_indices and cell.value:
                    cell.number_format = "DD-MMM-YY"

        light_fill = PatternFill("solid", fgColor="EBF3FB")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = light_fill

        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        status_col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "Status"), None)
        if status_col_idx:
            for row in ws.iter_rows(min_row=2):
                if row[status_col_idx - 1].value == "In Transit":
                    for cell in row:
                        cell.fill = yellow_fill

    buf.seek(0)
    return buf


# ── Per-vehicle orchestration ────────────────────────────────────────
def process_vehicle_type(vehicle_type, delhivery_df, master_bytes):
    all_sheets = pd.read_excel(io.BytesIO(master_bytes), sheet_name=None, dtype=object)
    
    master_df   = load_master(all_sheets, vehicle_type)
    pincode_map = load_pincode_map(all_sheets, vehicle_type)
    pincode_sheet = "Pin-Code" if vehicle_type == "2W" else "Pin-Codes"
    if pincode_sheet not in all_sheets:
        raise ReportError(f"The master file is missing the required '{pincode_sheet}' sheet.")
    pincode_df_full = all_sheets[pincode_sheet].copy()

    existing_lrs = set(master_df["Lr No. "].astype(str).str.strip())

    deliv_vt = delhivery_df[delhivery_df["vehicle_type"] == vehicle_type].copy()
    deliv_vt["LRN"] = normalize_lr(deliv_vt["LRN"])
    existing_deliv = deliv_vt[deliv_vt["LRN"].isin(existing_lrs)]
    new_deliv      = deliv_vt[~deliv_vt["LRN"].isin(existing_lrs)]

    master_df, updated = update_existing_rows(master_df, existing_deliv, "LRN", "Delivered Date")

    max_sno = master_df["S.NO"].max()
    existing_max_sno = int(max_sno) if (not master_df.empty and pd.notna(max_sno)) else 0
    new_rows = build_new_rows(new_deliv, pincode_map, vehicle_type, existing_max_sno)

    if new_rows:
        new_df = pd.DataFrame(new_rows)[MASTER_COLUMNS]
        master_df = pd.concat([master_df, new_df], ignore_index=True)

    out = save_master(master_df, pincode_df_full, vehicle_type)
    return out, {"updated": updated, "new": len(new_rows), "total": len(master_df)}


# ── Public entry point ───────────────────────────────────────────────
def generate(delhivery_files, file_2w, file_cv):
    """
    delhivery_files: list of uploaded CSV file objects.
    file_2w, file_cv: uploaded master .xlsx file objects.
    Returns dict: {'2W': (BytesIO, filename), 'CV': (BytesIO, filename), 'summary': {...}}.
    """
    frames = [clean_delhivery(f) for f in delhivery_files]
    delhivery_df = pd.concat(frames, ignore_index=True)
    delhivery_df["_rank"] = delhivery_df["Status_mapped"].apply(lambda x: 0 if x == "Delivered" else 1)
    delhivery_df = (delhivery_df.sort_values("_rank")
                    .drop_duplicates(subset="LRN", keep="first")
                    .drop(columns="_rank"))

    bytes_2w = file_2w.read()
    bytes_cv = file_cv.read()

    out_2w, sum_2w = process_vehicle_type("2W", delhivery_df, bytes_2w)
    out_cv, sum_cv = process_vehicle_type("CV", delhivery_df, bytes_cv)

    date_str = _report_date().strftime("%d%b%y")
    return {
        "2W": (out_2w, f"2W_{date_str}.xlsx"),
        "CV": (out_cv, f"CV_{date_str}.xlsx"),
        "summary": {
            "delhivery_rows": len(delhivery_df),
            "two_w": sum_2w,
            "cv": sum_cv,
        },
    }
