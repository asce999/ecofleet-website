import datetime
import openpyxl
import os
import logging
from openpyxl.utils import get_column_letter, column_index_from_string
from django.core.cache import cache
from core.workbook.helpers import atomic_save_workbook

HEADER_MAP = {
    'lr_number': ['LR NUMBER', 'LR No.', 'LR NO'],
    'pickup_date': ['Pickup Request Date', 'Pickup Date', 'Date'],
    'name': ['Name', 'Consignee Name', 'Customer Name', 'Consignee/Customer Name'],
    'address': ['Address', 'Consignee Address'],
    'contact_person': ['Contact Person'],
    'contact_number': ['Contact Number', 'Contact No.'],
    'city': ['City', 'Destination City'],
    'state': ['State', 'Consignee State'],
    'boxes': ['No Of Boxes', 'Boxes', 'No. of Boxes'],
    'weight_ef': ['Weight as per EcoFleet', 'Weight (EcoFleet)'],
    'weight_opt': ['Weight as per Optlog', 'Weight (Optlog)'],
    'status': ['Status', 'Status with Delhivery'],
    'delivered_on': ['Delivered on', 'Delivered on ', 'Actual Delivery Date', 'Delivery Date'],
    'tat': ['TAT'],
    'rate': ['Rate', 'Customer Rate'],
    'amount': ['Amount'],
    'vendor': ['Vendor', 'Vendor Name'],
    'vendor_rate': ['Vendor Rate'],
    'vendor_payment': ['CNG Paid & Vendor Payment', 'CNG Paid', 'Vendor Payment'],
}

def get_column_mapping(sheet):
    """
    Scans the first row of the sheet to map standard keys to 1-based column indices.
    """
    mapping = {}
    for col in range(1, 40): # Scan up to 40 columns
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        val_str = str(val).strip().lower()
        # Find which key in HEADER_MAP matches
        for key, variations in HEADER_MAP.items():
            if any(var.strip().lower() == val_str for var in variations):
                mapping[key] = col
                break
                
    # Also find 'SR. NO.' or '#' for serial number column
    for col in range(1, 40):
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        val_str = str(val).strip().lower()
        if val_str in ['sr. no.', '#', 'serial', 'sr no', 'sr.no.']:
            mapping['serial'] = col
            break
            
    return mapping

def find_totals_row(sheet, mapping):
    """
    Finds the first row containing total summaries or SUM formulas in the Amount column.
    """
    amount_col = mapping.get('amount', 17)
    for r in range(2, 250):  # Scan up to 250 rows
        val = sheet.cell(row=r, column=amount_col).value
        if isinstance(val, str) and 'SUM(' in val.upper():
            return r
        # Also check if Column 1 or Column 4 has total keywords
        col1_val = sheet.cell(row=r, column=1).value
        col4_val = sheet.cell(row=r, column=4).value
        for v in [col1_val, col4_val]:
            if isinstance(v, str) and any(x in v.upper() for x in ['TOTAL', 'GRAND TOTAL', 'JAMA', 'SUM', 'KHARCH']):
                return r
    return 64  # Default fallback if not found

def find_next_btpl_row(file_path, sheet_name='JUN 26'):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    lr_col = mapping.get('lr_number', 3)
    totals_row = find_totals_row(sheet, mapping)
    
    for r in range(2, totals_row):
        val = sheet.cell(row=r, column=lr_col).value
        if val is None or str(val).strip() == "":
            return r
    return None

def get_btpl_row_values(file_path, row, sheet_name='JUN 26'):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    
    def get_val(key):
        col = mapping.get(key)
        if col is None:
            return None
        return sheet.cell(row=row, column=col).value
        
    amount_val = get_val('amount')
    if isinstance(amount_val, str) and amount_val.startswith('='):
        amount_val = None
        
    return {
        'row_num': row,
        'lr_number': get_val('lr_number'),
        'pickup_date': get_val('pickup_date'),
        'name': get_val('name'),
        'address': get_val('address'),
        'contact_person': get_val('contact_person'),
        'contact_number': get_val('contact_number'),
        'city': get_val('city'),
        'state': get_val('state'),
        'boxes': get_val('boxes'),
        'weight_ef': get_val('weight_ef'),
        'weight_opt': get_val('weight_opt'),
        'status': get_val('status'),
        'delivered_on': get_val('delivered_on'),
        'tat': get_val('tat'),
        'rate': get_val('rate'),
        'amount': amount_val,
        'vendor': get_val('vendor'),
        'vendor_rate': get_val('vendor_rate'),
        'vendor_payment': get_val('vendor_payment'),
    }

def add_btpl_shipment(file_path, row_data, sheet_name='JUN 26'):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    row = row_data['row_num']
    
    mapping = get_column_mapping(sheet)
    
    def clean(val):
        if val == "" or val is None:
            return None
        if isinstance(val, str):
            val_strip = val.strip()
            if val_strip == "":
                return None
            return val_strip
        return val

    def write_val(key, val):
        col = mapping.get(key)
        if col is not None:
            sheet.cell(row=row, column=col).value = clean(val)

    write_val('lr_number', row_data['lr_number'])
    
    # Auto-populate serial number if the column exists and is empty
    serial_col = mapping.get('serial')
    if serial_col is not None:
        current_serial = sheet.cell(row=row, column=serial_col).value
        if current_serial is None or str(current_serial).strip() == "":
            sheet.cell(row=row, column=serial_col).value = row - 1

    write_val('pickup_date', row_data['pickup_date'])
    write_val('name', row_data['name'])
    write_val('address', row_data['address'])
    write_val('contact_person', row_data['contact_person'])
    write_val('contact_number', row_data['contact_number'])
    write_val('city', row_data['city'])
    write_val('state', row_data['state'])
    write_val('boxes', row_data['boxes'])
    write_val('weight_ef', row_data['weight_ef'])
    write_val('weight_opt', row_data['weight_opt'])
    write_val('status', row_data['status'])
    write_val('delivered_on', row_data['delivered_on'])
    write_val('tat', row_data['tat'])
    write_val('rate', row_data['rate'])
    
    # Amount formula or manual override calculation
    amount_col = mapping.get('amount', 17)
    if amount_col is not None:
        manual_amount = clean(row_data.get('amount'))
        if manual_amount is not None:
            sheet.cell(row=row, column=amount_col).value = manual_amount
        else:
            rate_letter = get_column_letter(mapping.get('rate', 16))
            weight_letter = get_column_letter(mapping.get('weight_opt', 12))
            sheet.cell(row=row, column=amount_col).value = f"={rate_letter}{row}*{weight_letter}{row}"
        
    write_val('vendor', row_data['vendor'])
    write_val('vendor_rate', row_data['vendor_rate'])
    write_val('vendor_payment', row_data['vendor_payment'])
    
    atomic_save_workbook(wb, file_path)

def get_cell_value_by_ref(sheet, ref, memo):
    col_letter = ""
    row_num_str = ""
    for char in ref:
        if char.isalpha():
            col_letter += char
        elif char.isdigit():
            row_num_str += char
            
    try:
        col = column_index_from_string(col_letter)
        row = int(row_num_str)
        return evaluate_cell(sheet, row, col, memo)
    except Exception:
        return None

def get_cells_in_range(sheet, start_ref, end_ref, memo):
    start_col_letter = ""
    start_row_str = ""
    for char in start_ref:
        if char.isalpha():
            start_col_letter += char
        elif char.isdigit():
            start_row_str += char
            
    end_col_letter = ""
    end_row_str = ""
    for char in end_ref:
        if char.isalpha():
            end_col_letter += char
        elif char.isdigit():
            end_row_str += char
            
    try:
        start_col = column_index_from_string(start_col_letter)
        start_row = int(start_row_str)
        end_col = column_index_from_string(end_col_letter)
        end_row = int(end_row_str)
        
        vals = []
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                vals.append(evaluate_cell(sheet, r, c, memo))
        return vals
    except Exception:
        return []

def evaluate_cell(sheet, row, col, memo=None):
    if memo is None:
        memo = {}
    
    cell_id = (row, col)
    if cell_id in memo:
        return memo[cell_id]
        
    val = sheet.cell(row=row, column=col).value
    if val is None:
        memo[cell_id] = None
        return None
        
    if isinstance(val, str) and val.startswith('='):
        formula = val[1:].strip().upper()
        res = None
        try:
            if '*' in formula:
                parts = formula.split('*')
                if len(parts) == 2:
                    val1 = get_cell_value_by_ref(sheet, parts[0], memo)
                    val2 = get_cell_value_by_ref(sheet, parts[1], memo)
                    # Convert to numeric if possible
                    try:
                        res = float(val1 or 0) * float(val2 or 0)
                    except ValueError:
                        res = 0.0
            elif formula.startswith('SUM(') and formula.endswith(')'):
                range_str = formula[4:-1]
                if ':' in range_str:
                    start_ref, end_ref = range_str.split(':')
                    cells = get_cells_in_range(sheet, start_ref, end_ref, memo)
                    sum_val = 0.0
                    for c in cells:
                        if c is not None:
                            try:
                                sum_val += float(c)
                            except ValueError:
                                pass
                    res = sum_val
            elif '+' in formula:
                parts = formula.split('+')
                if len(parts) == 2:
                    val1 = get_cell_value_by_ref(sheet, parts[0], memo)
                    val2 = get_cell_value_by_ref(sheet, parts[1], memo)
                    try:
                        res = float(val1 or 0) + float(val2 or 0)
                    except ValueError:
                        res = 0.0
            elif '-' in formula:
                parts = formula.split('-')
                if len(parts) == 2:
                    val1 = get_cell_value_by_ref(sheet, parts[0], memo)
                    val2 = get_cell_value_by_ref(sheet, parts[1], memo)
                    try:
                        res = float(val1 or 0) - float(val2 or 0)
                    except ValueError:
                        res = 0.0
        except Exception as e:
            res = f"#ERROR: {e}"
            
        memo[cell_id] = res
        return res
    else:
        # Check if the static value looks like a float or int
        if isinstance(val, (int, float)):
            memo[cell_id] = val
            return val
        elif isinstance(val, str):
            try:
                if '.' in val:
                    fval = float(val)
                    memo[cell_id] = fval
                    return fval
                else:
                    ival = int(val)
                    memo[cell_id] = ival
                    return ival
            except ValueError:
                pass
        memo[cell_id] = val
        return val

def get_cached_btpl_raw_data(file_path, sheet_name):
    import os
    if not os.path.exists(file_path):
        return None
    mtime = os.path.getmtime(file_path)
    # Cache invalidation strategy:
    # We include `os.path.getmtime()` in the key. When the workbook is saved/edited,
    # the mtime changes, causing an immediate cache miss and re-evaluation.
    safe_sheet = sheet_name.replace(' ', '_')
    cache_key = f"btpl_raw_data_{safe_sheet}_{mtime}"
    cached = cache.get(cache_key)
    
    if cached:
        return cached

    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    totals_row = find_totals_row(sheet, mapping)
    memo = {}
    
    max_col = max(mapping.values()) if mapping else 20

    def format_cell(val):
        if val is None:
            return ""
        elif isinstance(val, (datetime.datetime, datetime.date)):
            return val.strftime('%d-%b-%y')
        elif isinstance(val, float):
            return str(int(val)) if val.is_integer() else f"{val:.2f}"
        else:
            return str(val)

    columns = []
    for c in range(1, max_col + 1):
        val = evaluate_cell(sheet, 1, c, memo)
        columns.append(format_cell(val))

    all_data_rows = []
    for r in range(2, sheet.max_row + 1):
        row_vals = []
        has_data = False
        for c in range(1, max_col + 1):
            val = evaluate_cell(sheet, r, c, memo)
            formatted = format_cell(val)
            row_vals.append(formatted)
            if formatted != "":
                has_data = True
        if has_data:
            all_data_rows.append({'row_num': r, 'cells': row_vals})
            
    lr_col = mapping.get('lr_number', 3)
    auto_next_row = None
    row_values_map = {}
    
    for r in range(2, totals_row):
        lr_val = sheet.cell(row=r, column=lr_col).value
        if auto_next_row is None and (lr_val is None or str(lr_val).strip() == ""):
            auto_next_row = r
            
        def get_val(key, row=r):
            col = mapping.get(key)
            if col is None:
                return None
            return sheet.cell(row=row, column=col).value
            
        amount_val = get_val('amount')
        if isinstance(amount_val, str) and amount_val.startswith('='):
            amount_val = None
            
        row_values_map[r] = {
            'row_num': r,
            'lr_number': get_val('lr_number'),
            'pickup_date': get_val('pickup_date'),
            'name': get_val('name'),
            'address': get_val('address'),
            'contact_person': get_val('contact_person'),
            'contact_number': get_val('contact_number'),
            'city': get_val('city'),
            'state': get_val('state'),
            'boxes': get_val('boxes'),
            'weight_ef': get_val('weight_ef'),
            'weight_opt': get_val('weight_opt'),
            'status': get_val('status'),
            'delivered_on': get_val('delivered_on'),
            'tat': get_val('tat'),
            'rate': get_val('rate'),
            'amount': amount_val,
            'vendor': get_val('vendor'),
            'vendor_rate': get_val('vendor_rate'),
            'vendor_payment': get_val('vendor_payment'),
        }

    cached = {
        'mapping': mapping,
        'totals_row': totals_row,
        'auto_next_row': auto_next_row,
        'columns': columns,
        'all_data_rows': all_data_rows,
        'row_values_map': row_values_map,
    }
    
    # Cache for 1 hour. It will invalidate early if the file changes due to mtime in the key.
    cache.set(cache_key, cached, timeout=3600)
    return cached

def get_btpl_preview(file_path, sheet_name='JUN 26', page=1, page_size=20, sheet=None, mapping=None, memo=None):
    """Return paginated preview data, reading from cache if possible."""
    cached = get_cached_btpl_raw_data(file_path, sheet_name)
    if not cached:
        return {'columns': [], 'rows': [], 'total_rows': 0, 'page': 1, 'page_size': page_size, 'total_pages': 0}
        
    columns = cached['columns']
    all_data_rows = cached['all_data_rows']
    
    total_rows = len(all_data_rows)
    total_pages = max(1, (total_rows + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))

    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_rows = all_data_rows[start_idx:end_idx]

    return {
        'columns': columns,
        'rows': page_rows,
        'total_rows': total_rows,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }

def get_btpl_page_data(file_path, sheet_name='JUN 26', target_row=None, page=1, page_size=20):
    """
    Returns all data needed for the BTPL page (mapping, totals_row, next_row, row_values, preview).
    Reads instantly from cache to avoid openpyxl parsing.
    """
    cached = get_cached_btpl_raw_data(file_path, sheet_name)
    if not cached:
        return None
        
    mapping = cached['mapping']
    totals_row = cached['totals_row']
    auto_next_row = cached['auto_next_row']
    
    if target_row and 2 <= target_row < totals_row:
        next_row = target_row
    else:
        next_row = auto_next_row
        
    row_values = {}
    if next_row:
        row_values = cached['row_values_map'].get(next_row, {})
        
    preview = get_btpl_preview(file_path, sheet_name=sheet_name, page=page, page_size=page_size)
    
    return {
        'mapping': mapping,
        'totals_row': totals_row,
        'next_row': next_row,
        'row_values': row_values,
        'preview': preview,
    }

def clear_btpl_row(file_path, row, sheet_name='JUN 26'):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    mapping = get_column_mapping(sheet)
    
    # Clear all cells in the row except serial number and amount formula
    for key, col in mapping.items():
        if key not in ['serial', 'amount']:
            sheet.cell(row=row, column=col).value = None
            
    # Re-apply Amount formula to avoid empty calculation
    rate_letter = get_column_letter(mapping.get('rate', 16))
    weight_letter = get_column_letter(mapping.get('weight_opt', 12))
    amount_col = mapping.get('amount', 17)
    if amount_col is not None:
        sheet.cell(row=row, column=amount_col).value = f"={rate_letter}{row}*{weight_letter}{row}"
        
    atomic_save_workbook(wb, file_path)
