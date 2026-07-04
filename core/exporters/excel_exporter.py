import openpyxl
from io import BytesIO
from core.models import Shipment

class ExcelExporter:
    # TODO(phase-3): wired when dual-read/export lands
    def export_ftl_shipments(self):
        """Generates an FTL Excel file directly from the PostgreSQL Shipment domain models."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = [
            "Date of Booking", "ETD", "Actual Delivery Date", 
            "Consignor", "From Location", "Consignee", 
            "LR Number", "To Location", "Vehicle Number", "Vendor"
        ]
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_title
            
        shipments = Shipment.objects.filter(shipment_type='FTL').select_related('vehicle').order_by('-dispatch_date')
        
        for row_num, shipment in enumerate(shipments, 2):
            ws.cell(row=row_num, column=1).value = shipment.dispatch_date
            
            if shipment.actual_eta:
                ws.cell(row=row_num, column=3).value = shipment.actual_eta.date()
            if shipment.expected_eta:
                ws.cell(row=row_num, column=2).value = shipment.expected_eta.date()
            
            ws.cell(row=row_num, column=4).value = shipment.metadata.get('consignor', '')
            ws.cell(row=row_num, column=5).value = shipment.origin
            ws.cell(row=row_num, column=6).value = shipment.metadata.get('consignee', '')
            ws.cell(row=row_num, column=7).value = shipment.metadata.get('lr_number', '')
            ws.cell(row=row_num, column=8).value = shipment.destination
            ws.cell(row=row_num, column=9).value = shipment.vehicle.registration_number if shipment.vehicle else ''
            ws.cell(row=row_num, column=10).value = shipment.metadata.get('vendor', '')
            
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
