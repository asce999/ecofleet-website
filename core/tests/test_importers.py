import os
import openpyxl
import tempfile
from django.test import TestCase
from core.models import ImportJob, Shipment, Vehicle
from core.importers.excel_importer import ExcelImporter

class ImporterTests(TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.file_path = os.path.join(self.tmp_dir.name, "dummy_ftl.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        
        headers = ["Date of Booking", "LR Number", "Vehicle Number", "Origin", "Destination"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        ws.cell(row=2, column=1, value="2026-06-25")
        ws.cell(row=2, column=2, value="LR1001")
        ws.cell(row=2, column=3, value="MH12AB1234")
        ws.cell(row=2, column=4, value="Pune")
        ws.cell(row=2, column=5, value="Mumbai")

        wb.save(self.file_path)

    def tearDown(self):
        self.tmp_dir.cleanup()

    def test_excel_importer_creates_shipments(self):
        job = ImportJob.objects.create(workbook_type='FTL')
        importer = ExcelImporter()
        
        importer.process_ftl_workbook(job.id, self.file_path)
        
        job.refresh_from_db()
        self.assertEqual(job.status, 'COMPLETED')
        self.assertEqual(job.processed_rows, 1)
        self.assertEqual(job.failed_rows, 0)
        
        self.assertEqual(Shipment.objects.count(), 1)
        shipment = Shipment.objects.first()
        self.assertEqual(shipment.shipment_type, 'FTL')
        self.assertEqual(shipment.origin, 'Pune')
        self.assertEqual(shipment.destination, 'Mumbai')
        
        self.assertEqual(Vehicle.objects.count(), 1)
        vehicle = Vehicle.objects.first()
        self.assertEqual(vehicle.registration_number, 'MH12AB1234')
