import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_salary_export(sheet_name: str, salary_data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary_{sheet_name}"

    # Add the top header rows to match BTPL
    ws.cell(row=1, column=2, value="CV DISTRIBUTIONS")
    ws.cell(row=2, column=2, value="BARAMATI TRADE PVT LTD")
    ws.cell(row=3, column=2, value="D-16 MIDC AREA BARAMATI,DIST-PUNE")

    headers = [
        (2, "Sr No"), (3, "Department"), (4, "EmpID"), (5, "ESIC NO"), (6, "PF NO."), (7, "UAN NO"), 
        (8, "Name of the  Employee"), (9, "Aadhar Card no"), (10, "DOB"), (11, "Doj"), (12, "Mobile No"), 
        (13, "Cosmos Account No"), (14, "Payble Days"), (15, "Extra Days"), (16, "BASIC=    492.12"), 
        (17, "Adhoc Salary Increase"), (18, "SP.ALLOWANCE = 96.58"), (19, "ADHOC ALLOWANCE"), (20, "TOTAL (A)"), 
        (21, "OTHER ALLOWANCE"), (22, "H.R.A. 5% BASIC+SPL ALL"), (23, "Leave Payment"), (24, "EXTRA PAYMENT"), 
        (25, "TOTAL (B)"), (26, "P.F.12% ON Basic +Spl All.Total(A)"), (27, "Esic 0.75% Total "), (28, "P.Tax On Total (B)"), 
        (29, "Advance"), (30, "LWF"), (31, "Canteen"), (32, "Other"), (33, "Sub Total"), (34, "Payment")
    ]
    
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border_side = Side(border_style='thin', color='E5E7EB')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center')

    for col_idx, header_title in headers:
        cell = ws.cell(row=4, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    for i, emp in enumerate(salary_data['employees']):
        r = i + 5
        
        def write_cell(col, val, fmt=None):
            c = ws.cell(row=r, column=col, value=val)
            c.border = thin_border
            if fmt:
                c.number_format = fmt
            return c

        write_cell(2, i + 1)
        write_cell(3, emp.get('department', ''))
        write_cell(4, emp.get('emp_id', ''))
        write_cell(5, emp.get('esic_no', ''))
        write_cell(6, emp.get('pf_no', ''))
        write_cell(7, emp.get('uan_no', ''))
        write_cell(8, emp.get('name', ''))
        write_cell(9, '')
        write_cell(10, '')
        write_cell(11, '')
        write_cell(12, emp.get('phone', ''))
        write_cell(13, '')
        write_cell(14, float(emp.get('payable_days', 0)))
        write_cell(15, float(emp.get('extra_days', 0)))
        
        write_cell(16, float(emp.get('basic', 0)), '#,##0.00')
        write_cell(17, float(emp.get('adhoc_salary_increase', 0)), '#,##0.00')
        write_cell(18, float(emp.get('sp_allowance', 0)), '#,##0.00')
        write_cell(19, float(emp.get('adhoc_allowance', 0)), '#,##0.00')
        write_cell(20, float(emp.get('total_a', 0)), '#,##0.00')
        
        write_cell(21, float(emp.get('other_allowance', 0)), '#,##0.00')
        write_cell(22, float(emp.get('hra', 0)), '#,##0.00')
        write_cell(23, float(emp.get('leave_payment', 0)), '#,##0.00')
        write_cell(24, float(emp.get('extra_payment', 0)), '#,##0.00')
        c = write_cell(25, float(emp.get('total_b', 0)), '#,##0.00')
        c.font = Font(bold=True)
        
        write_cell(26, float(emp.get('pf', 0)), '#,##0.00')
        write_cell(27, float(emp.get('esic_employee', 0)), '#,##0.00')
        write_cell(28, float(emp.get('pt', 0)), '#,##0.00')
        write_cell(29, float(emp.get('advance', 0)), '#,##0.00')
        write_cell(30, float(emp.get('lwf', 0)), '#,##0.00')
        write_cell(31, float(emp.get('canteen', 0)), '#,##0.00')
        write_cell(32, float(emp.get('other_deduction', 0)), '#,##0.00')
        c = write_cell(33, float(emp.get('sub_total_deductions', 0)), '#,##0.00')
        c.font = Font(color='FF0000')
        
        c = write_cell(34, float(emp.get('net_payment', 0)), '#,##0.00')
        c.font = Font(bold=True, color='008000')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
