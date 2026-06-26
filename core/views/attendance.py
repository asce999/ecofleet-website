from django.views.decorators.cache import never_cache
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404
from core.models import AttendanceWorkbook
from core.forms import AttendanceWorkbookUploadForm
from core.decorators import staff_required, tool_permission_required
from core import attendance as attendance_logic
import os
from core.models import ToolRun
from django.contrib import messages
from django.conf import settings
from django.urls import reverse
import logging

logger = logging.getLogger('core')

@staff_required
@tool_permission_required('attendance')
@never_cache
def attendance_sheet(request):
    from django.urls import reverse
    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    
    if not file_path or not os.path.exists(file_path):
        return render(request, 'core/portal/attendance_form.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    if request.method == 'POST':
        try:
            modified = attendance_logic.save_attendance(file_path, sheet_name, request.POST)
            if modified:
                ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_ATTENDANCE,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"Update: {sheet_name}",
                    detail=f"Updated attendance sheet '{sheet_name}'"
                )
                messages.success(request, "Attendance updated successfully.")
                return redirect(reverse('attendance_sheet') + '?success=1')
            else:
                messages.info(request, "No changes were made to the attendance sheet.")
                return redirect('attendance_sheet')
        except Exception as e:
            messages.error(request, f"Failed to save attendance: {e}")
            return redirect('attendance_sheet')
            
    manual_year = request.POST.get('manual_year') or request.GET.get('manual_year')
    manual_month = request.POST.get('manual_month') or request.GET.get('manual_month')
    
    # Load sheet data
    data = attendance_logic.get_attendance_data(file_path, sheet_name, manual_year, manual_month)
    if not data:
        return render(request, 'core/portal/attendance_form.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    if data.get('requires_manual_date'):
        return render(request, 'core/portal/attendance_form.html', {
            'active': 'attendance',
            'requires_manual_date': True,
            'sheet_name': sheet_name,
            'wb': wb_obj,
            'sheet_names': data.get('sheet_names', [])
        })
        
    days_list = list(range(1, data['days_in_month'] + 1))
    days_info_list = []
    for d in days_list:
        days_info_list.append({
            'day': d,
            'is_holiday': data['day_headers'][d]['is_holiday']
        })
        
    employees_list = []
    for emp in data['employees']:
        days_attendance = []
        for d in days_list:
            days_attendance.append({
                'day': d,
                'status': emp['attendance'].get(d, ''),
                'is_holiday': data['day_headers'][d]['is_holiday']
            })
        employees_list.append({
            'row_idx': emp['row_idx'],
            'name': emp['name'],
            'phone': emp['phone'],
            'days_attendance': days_attendance
        })
        
    context = {
        'active': 'attendance',
        'sheet_name': data['sheet_name'],
        'days_in_month': data['days_in_month'],
        'days_info_list': days_info_list,
        'employees': employees_list,
        'sheet_names': data['sheet_names'],
        'wb': wb_obj,
        'success': request.GET.get('success') == '1',
        'manual_year': data.get('year'),
        'manual_month': data.get('month'),
    }
    return render(request, 'core/portal/attendance_form.html', context)


@staff_required
@tool_permission_required('attendance')
def attendance_download(request):
    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    if not file_path or not os.path.exists(file_path):
        raise Http404("Attendance workbook file not found.")
        
    filename = wb_obj.original_name if wb_obj else "Attendance_Sheet.xlsx"
    response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_required
@tool_permission_required('attendance')
def attendance_settings(request):
    """Manage active Attendance workbook and active sheet tab."""
    wb_obj, file_path, current_sheet_name = attendance_logic.get_active_attendance_workbook()
    
    # Read sheets from Excel if file_path exists
    sheets = []
    if file_path and os.path.exists(file_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
        except Exception:
            sheets = ['JUNE 2026']
    else:
        sheets = ['JUNE 2026']
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'remove':
            AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
            if AttendanceWorkbook.objects.count() == 0:
                AttendanceWorkbook.objects.create(
                    original_name='Attendance_Sheet.xlsx',
                    active_sheet='JUNE 2026',
                    uploaded_by=request.user,
                    is_active=False
                )
            messages.success(request, "Attendance sheet removed entirely. Portal is now in empty state.")
            return redirect('attendance_settings')
            
        elif action == 'load_default':
            import shutil
            root_file_path = os.path.join(settings.BASE_DIR, 'Attendance_Sheet.xlsx')
            target_dir = os.path.join(settings.MEDIA_ROOT, 'attendance')
            os.makedirs(target_dir, exist_ok=True)
            dest_path = os.path.join(target_dir, 'Attendance_Sheet.xlsx')
            shutil.copy2(root_file_path, dest_path)
            
            AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
            
            wb_obj = AttendanceWorkbook.objects.create(
                original_name='Attendance_Sheet.xlsx',
                active_sheet='JUNE 2026',
                uploaded_by=request.user,
                is_active=True
            )
            wb_obj.file.name = 'attendance/Attendance_Sheet.xlsx'
            wb_obj.save(update_fields=['file'])
            
            try:
                wb = openpyxl.load_workbook(dest_path, read_only=True)
                if wb.sheetnames:
                    wb_obj.active_sheet = wb.sheetnames[-1]
                    wb_obj.save(update_fields=['active_sheet'])
            except Exception:
                pass
                
            messages.success(request, "Default attendance workbook loaded successfully.")
            return redirect('attendance_settings')
            
        elif action == 'change_sheet':
            form = AttendanceWorkbookUploadForm(request.POST, sheets=sheets)
            if form.is_valid():
                sheet_sel = form.cleaned_data.get('active_sheet')
                if wb_obj:
                    wb_obj.active_sheet = sheet_sel
                    wb_obj.save(update_fields=['active_sheet'])
                else:
                    import shutil
                    root_file_path = os.path.join(settings.BASE_DIR, 'Attendance_Sheet.xlsx')
                    target_dir = os.path.join(settings.MEDIA_ROOT, 'attendance')
                    os.makedirs(target_dir, exist_ok=True)
                    dest_path = os.path.join(target_dir, 'Attendance_Sheet.xlsx')
                    shutil.copy2(root_file_path, dest_path)
                    
                    wb_obj = AttendanceWorkbook.objects.create(
                        original_name='Attendance_Sheet.xlsx',
                        active_sheet=sheet_sel,
                        uploaded_by=request.user,
                        is_active=True
                    )
                    wb_obj.file.name = 'attendance/Attendance_Sheet.xlsx'
                    wb_obj.save(update_fields=['file'])
                    
                messages.success(request, f"Active sheet tab changed to '{sheet_sel}'.")
                return redirect('attendance_sheet')
                
        elif action == 'upload':
            form = AttendanceWorkbookUploadForm(request.POST, request.FILES, sheets=sheets)
            if form.is_valid():
                if upload:
                    logger.info(f"Attendance workbook upload started: {upload.name}")
                    new_wb = AttendanceWorkbook.objects.create(
                        file=upload, original_name=upload.name,
                        uploaded_by=request.user, is_active=False
                    )
                    
                    try:
                        import openpyxl
                        temp_wb = openpyxl.load_workbook(new_wb.file.path, read_only=True)
                        uploaded_sheets = temp_wb.sheetnames
                        default_sheet = uploaded_sheets[-1] if uploaded_sheets else 'JUNE 2026'
                    except Exception as e:
                        logger.error(f"Excel parsing failed for {upload.name}: {e}")
                        default_sheet = 'JUNE 2026'
                        
                    new_wb.active_sheet = default_sheet
                    new_wb.save(update_fields=['active_sheet'])
                    
                    AttendanceWorkbook.objects.filter(is_active=True).update(is_active=False)
                    new_wb.is_active = True
                    new_wb.save(update_fields=['is_active'])
                    
                    logger.info(f"Attendance workbook upload completed: {upload.name}")
                    messages.success(request, f"Uploaded workbook '{upload.name}' is now the active attendance workbook.")
                    return redirect('attendance_settings')
                else:
                    logger.warning("Attendance file upload failed: No file provided.")
                    messages.error(request, "Please choose a valid .xlsx file to upload.")
                    
        elif action == 'update_salary_config':
            from core.models import SalaryConfig
            from core.forms import SalaryConfigForm
            config = SalaryConfig.get_solo()
            salary_form = SalaryConfigForm(request.POST, instance=config)
            if salary_form.is_valid():
                salary_form.save()
                messages.success(request, "Salary configuration updated successfully.")
                return redirect('attendance_settings')
    else:
        form = AttendanceWorkbookUploadForm(initial={'active_sheet': current_sheet_name}, sheets=sheets)
        
    from core.models import SalaryConfig
    from core.forms import SalaryConfigForm
    salary_form = SalaryConfigForm(instance=SalaryConfig.get_solo())
        
    context = {
        'active': 'attendance',
        'form': form,
        'salary_form': salary_form,
        'wb': wb_obj,
        'current_sheet': current_sheet_name,
        'sheets': sheets,
    }
    return render(request, 'core/portal/attendance_settings.html', context)



@staff_required
@tool_permission_required('attendance')
@never_cache
def salary_calculator(request):
    from core.models import EmployeeSalaryOverride, SalaryConfig
    from decimal import Decimal

    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    
    if not file_path or not os.path.exists(file_path):
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'no_sheet': True,
        })

    manual_year = request.POST.get('manual_year') or request.GET.get('manual_year')
    manual_month = request.POST.get('manual_month') or request.GET.get('manual_month')
    
    data = attendance_logic.get_attendance_data(file_path, sheet_name, manual_year, manual_month)
    if not data:
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    if data.get('requires_manual_date'):
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'requires_manual_date': True,
            'sheet_name': sheet_name,
            'wb': wb_obj,
            'sheet_names': data.get('sheet_names', [])
        })
        
    if request.method == 'POST':
        # Save salaries
        for emp in data['employees']:
            emp_name = emp['name']
            inc_val = request.POST.get(f'override_inc_{emp_name}')
            allow_val = request.POST.get(f'override_allow_{emp_name}')
            adv_val = request.POST.get(f'override_adv_{emp_name}')
            lwf_val = request.POST.get(f'override_lwf_{emp_name}')
            oth_val = request.POST.get(f'override_other_{emp_name}')
            csh_val = request.POST.get(f'override_cash_{emp_name}')
            
            if any(v is not None for v in [inc_val, allow_val, adv_val, lwf_val, oth_val, csh_val]):
                try:
                    def parse_dec(val):
                        if val is None or val.strip() == '': return Decimal('0.00')
                        return Decimal(val.strip())
                        
                    EmployeeSalaryOverride.objects.update_or_create(
                        employee_name=emp_name,
                        defaults={
                            'adhoc_salary_increase_pct': parse_dec(inc_val),
                            'adhoc_allowance_monthly_amount': parse_dec(allow_val),
                            'advance': parse_dec(adv_val),
                            'lwf': parse_dec(lwf_val),
                            'other_deduction': parse_dec(oth_val),
                            'cash_payment': parse_dec(csh_val)
                        }
                    )
                except Exception as e:
                    logger.error(f"Validation failure for salary overrides of {emp_name}: {e}")
                    messages.error(request, f"Invalid override values for {emp_name}: {e}")
                    
        logger.info(f"Salary calculation completed for sheet: {sheet_name}")
        messages.success(request, "Salaries saved successfully.")
        manual_year = request.POST.get('manual_year')
        manual_month = request.POST.get('manual_month')
        redirect_url = reverse('salary_calculator')
        if manual_year and manual_month:
            redirect_url += f"?manual_year={manual_year}&manual_month={manual_month}"
        return redirect(redirect_url)

    # Ensure all employees in the sheet have an EmployeeSalaryOverride record
    employee_names = [emp['name'] for emp in data['employees']]
    existing_salaries = EmployeeSalaryOverride.objects.filter(employee_name__in=employee_names).values_list('employee_name', flat=True)
    for name in employee_names:
        if name not in existing_salaries:
            EmployeeSalaryOverride.objects.get_or_create(employee_name=name)

    salary_data = attendance_logic.calculate_salary_data(data)
    
    context = {
        'active': 'attendance',
        'sheet_name': sheet_name,
        'wb': wb_obj,
        'salary_data': salary_data,
        'sheet_names': data['sheet_names'],
        'manual_year': data.get('year'),
        'manual_month': data.get('month'),
    }
    return render(request, 'core/portal/attendance_salary.html', context)


@staff_required
@tool_permission_required('attendance')
def salary_calculator_export(request):
    import io
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    if not file_path or not os.path.exists(file_path):
        messages.error(request, "Attendance workbook file not found.")
        return redirect('salary_calculator')

    manual_year = request.GET.get('manual_year')
    manual_month = request.GET.get('manual_month')
    
    data = attendance_logic.get_attendance_data(file_path, sheet_name, manual_year, manual_month)
    if not data or data.get('requires_manual_date'):
        messages.error(request, "Cannot export: Invalid sheet date or data.")
        return redirect('salary_calculator')

    salary_data = attendance_logic.calculate_salary_data(data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary_{sheet_name}"

    # Add the top header rows to match BTPL
    ws.cell(row=1, column=2, value="CV DISTRIBUTIONS")
    ws.cell(row=2, column=2, value="BARAMATI TRADE PVT LTD")
    ws.cell(row=3, column=2, value="D-16 MIDC AREA BARAMATI,DIST-PUNE")

    headers = [
        (2, "Sr No"), (3, "Department"), (4, "EmpID"), (5, "ESIC NO"), (6, "PF NO."), (7, "UAN NO"), 
        (8, "Name of the  Employee"), (9, "Aadhar Card no"), (10, "DOB"), (11, "Doj"), (12, "Mobile No"), 
        (13, "Cosmos Account No"), (14, "Payble Days"), (15, "Extra Days"), (16, "BASIC=    492.12"), 
        (17, "Adhoc Salary Increase"), (18, "SP.ALLOWANCE = 96.58"), (19, "ADHOC ALLOWANCE"), (20, "TOTAL (A)"), 
        (21, "OTHER ALLOWANCE"), (22, "H.R.A. 5% BASIC+SPL ALL"), (23, "Leave Payment"), (24, "EXTRA PAYMENT"), 
        (25, "TOTAL (B)"), (26, "P.F.12% ON Basic +Spl All.Total(A)"), (27, "Esic 0.75% Total "), (28, "P.Tax On Total (B)"), 
        (29, "Advance"), (30, "LWF"), (31, "Canteen"), (32, "Other"), (33, "Sub Total"), (34, "Payment")
    ]
    
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border_side = Side(border_style='thin', color='E5E7EB')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center')

    for col_idx, header_title in headers:
        cell = ws.cell(row=4, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    for i, emp in enumerate(salary_data['employees']):
        r = i + 5
        
        def write_cell(col, val, fmt=None):
            c = ws.cell(row=r, column=col, value=val)
            c.border = thin_border
            if fmt:
                c.number_format = fmt
            return c

        write_cell(2, i + 1)
        write_cell(3, emp.get('department', ''))
        write_cell(4, emp.get('emp_id', ''))
        write_cell(5, emp.get('esic_no', ''))
        write_cell(6, emp.get('pf_no', ''))
        write_cell(7, emp.get('uan_no', ''))
        write_cell(8, emp.get('name', ''))
        write_cell(9, '')
        write_cell(10, '')
        write_cell(11, '')
        write_cell(12, emp.get('phone', ''))
        write_cell(13, '')
        write_cell(14, float(emp.get('payable_days', 0)))
        write_cell(15, float(emp.get('extra_days', 0)))
        
        write_cell(16, float(emp.get('basic', 0)), '#,##0.00')
        write_cell(17, float(emp.get('adhoc_salary_increase', 0)), '#,##0.00')
        write_cell(18, float(emp.get('sp_allowance', 0)), '#,##0.00')
        write_cell(19, float(emp.get('adhoc_allowance', 0)), '#,##0.00')
        write_cell(20, float(emp.get('total_a', 0)), '#,##0.00')
        
        write_cell(21, float(emp.get('other_allowance', 0)), '#,##0.00')
        write_cell(22, float(emp.get('hra', 0)), '#,##0.00')
        write_cell(23, float(emp.get('leave_payment', 0)), '#,##0.00')
        write_cell(24, float(emp.get('extra_payment', 0)), '#,##0.00')
        c = write_cell(25, float(emp.get('total_b', 0)), '#,##0.00')
        c.font = Font(bold=True)
        
        write_cell(26, float(emp.get('pf', 0)), '#,##0.00')
        write_cell(27, float(emp.get('esic_employee', 0)), '#,##0.00')
        write_cell(28, float(emp.get('pt', 0)), '#,##0.00')
        write_cell(29, float(emp.get('advance', 0)), '#,##0.00')
        write_cell(30, float(emp.get('lwf', 0)), '#,##0.00')
        write_cell(31, float(emp.get('canteen', 0)), '#,##0.00')
        write_cell(32, float(emp.get('other_deduction', 0)), '#,##0.00')
        c = write_cell(33, float(emp.get('sub_total_deductions', 0)), '#,##0.00')
        c.font = Font(color='FF0000')
        
        c = write_cell(34, float(emp.get('net_payment', 0)), '#,##0.00')
        c.font = Font(bold=True, color='008000')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Salary_{sheet_name}.xlsx"'
    return response
