from django.views.decorators.cache import never_cache
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404, JsonResponse
from core.models import BtplWorkbook
from core.forms import BtplShipmentForm, BtplWorkbookUploadForm
from core.decorators import staff_required, tool_permission_required, tool_permission_required
from core import btpl as btpl_logic
import os
import json


def get_active_btpl_workbook():
    from .models import BtplWorkbook
    # If there are BtplWorkbook records but no active one, it means sheet was removed entirely!
    wb_count = BtplWorkbook.objects.count()
    if wb_count > 0:
        wb_obj = BtplWorkbook.active()
        if not wb_obj:
            return None, None, None
        return wb_obj, wb_obj.file.path, wb_obj.active_sheet
    else:
        # Initial fresh state: fall back to default
        file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
        sheet_name = 'JUN 26'
        return None, file_path, sheet_name


@staff_required
@tool_permission_required('btpl')
@never_cache
def btpl_sheet(request):
    from django.urls import reverse
    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    
    if not file_path:
        return render(request, 'core/portal/btpl_form.html', {
            'active': 'btpl',
            'no_sheet': True,
        })

    page = int(request.GET.get('page', 1))
    page_data = btpl_logic.get_btpl_page_data(
        file_path, sheet_name=sheet_name, page=page, page_size=20
    )

    if page_data is None:
        return render(request, 'core/portal/btpl_form.html', {
            'active': 'btpl',
            'no_sheet': True,
        })

    totals_row = page_data['totals_row']
    next_row = page_data['next_row']

    context = {
        'active': 'btpl',
        'preview': page_data['preview'],
        'next_row': next_row,
        'totals_row': totals_row,
        'max_shipment_row': totals_row - 1,
        'sheet_name': sheet_name,
        'wb': wb_obj,
    }
    return render(request, 'core/portal/btpl_form.html', context)


@staff_required
@tool_permission_required('btpl')
@never_cache
def btpl_api(request):
    """JSON API endpoint for AJAX BTPL operations."""
    import json
    from django.http import JsonResponse

    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    if not file_path:
        return JsonResponse({'error': 'No active workbook'}, status=404)

    action = request.GET.get('action') or request.POST.get('action', '')

    # GET: fetch row data for editing
    if action == 'get_row':
        row_num = int(request.GET.get('row', 0))
        if row_num < 2:
            return JsonResponse({'error': 'Invalid row'}, status=400)
        row_data = btpl_logic.get_btpl_row_values(file_path, row_num, sheet_name=sheet_name)
        # Convert dates to strings for JSON
        for key in ['pickup_date', 'delivered_on']:
            val = row_data.get(key)
            if isinstance(val, (datetime.datetime, datetime.date)):
                row_data[key] = val.strftime('%Y-%m-%d')
        return JsonResponse({'row_data': row_data})

    # GET: paginated preview
    if action == 'preview':
        page = int(request.GET.get('page', 1))
        preview = btpl_logic.get_btpl_preview(file_path, sheet_name=sheet_name, page=page, page_size=20)
        return JsonResponse({'preview': preview})

    # POST: save row
    if action == 'save' and request.method == 'POST':
        form = BtplShipmentForm(request.POST)
        if form.is_valid():
            row_data = form.cleaned_data
            target_row = row_data['row_num']
            try:
                btpl_logic.add_btpl_shipment(file_path, row_data, sheet_name=sheet_name)
                run = ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_BTPL,
                    status=ToolRun.STATUS_SUCCESS,
                    reference=f"LR: {row_data.get('lr_number') or ''}",
                    detail=f"Row: {target_row} · Sheet: {sheet_name} · Consignee: {row_data.get('name') or ''}"
                )
                return JsonResponse({'success': True, 'run_id': run.pk, 'row': target_row})
            except Exception as e:
                ToolRun.objects.create(
                    user=request.user,
                    tool=ToolRun.TOOL_BTPL,
                    status=ToolRun.STATUS_FAILED,
                    detail=f"Error writing row {target_row}: {e}"
                )
                return JsonResponse({'error': str(e)}, status=500)
        else:
            errors = {k: v[0] for k, v in form.errors.items()}
            return JsonResponse({'error': 'Validation failed', 'field_errors': errors}, status=400)

    # POST: delete row
    if action == 'delete' and request.method == 'POST':
        row_num = int(request.POST.get('row', 0))
        # Get totals_row to validate
        page_data = btpl_logic.get_btpl_page_data(file_path, sheet_name=sheet_name, page=1, page_size=1)
        totals_row = page_data['totals_row'] if page_data else 64
        if row_num < 2 or row_num >= totals_row:
            return JsonResponse({'error': 'Invalid row number'}, status=400)
        try:
            btpl_logic.clear_btpl_row(file_path, row_num, sheet_name=sheet_name)
            ToolRun.objects.create(
                user=request.user,
                tool=ToolRun.TOOL_BTPL,
                status=ToolRun.STATUS_SUCCESS,
                reference=f"Clear Row: {row_num}",
                detail=f"Cleared all shipment details in Row {row_num} on sheet {sheet_name}"
            )
            return JsonResponse({'success': True, 'row': row_num})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # GET: next available row
    if action == 'next_row':
        next_row = btpl_logic.find_next_btpl_row(file_path, sheet_name=sheet_name)
        return JsonResponse({'next_row': next_row})

    return JsonResponse({'error': 'Unknown action'}, status=400)


@staff_required
@tool_permission_required('btpl')
def btpl_download(request):
    wb_obj, file_path, sheet_name = get_active_btpl_workbook()
    if not file_path or not os.path.exists(file_path):
        raise Http404("BTPL Shipments file not found.")
    
    filename = wb_obj.original_name if wb_obj else "BTPL_Shipments.xlsx"
    response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_required
@tool_permission_required('btpl')
def btpl_settings(request):
    """Manage active BTPL shipment workbook and active sheet tab."""
    wb_obj, file_path, current_sheet_name = get_active_btpl_workbook()
    
    # Read sheets from Excel if file_path exists
    sheets = []
    if file_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
        except Exception:
            sheets = ['JUN 26']
    else:
        sheets = ['JUN 26']
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'remove':
            BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
            if BtplWorkbook.objects.count() == 0:
                BtplWorkbook.objects.create(
                    original_name='BTPL_Shipments.xlsx',
                    active_sheet='JUN 26',
                    uploaded_by=request.user,
                    is_active=False
                )
            messages.success(request, "BTPL shipment sheet removed entirely. Portal is now in empty state.")
            return redirect('btpl_settings')
            
        elif action == 'load_default':
            import shutil
            root_file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
            target_dir = os.path.join(settings.MEDIA_ROOT, 'btpl')
            os.makedirs(target_dir, exist_ok=True)
            dest_path = os.path.join(target_dir, 'BTPL_Shipments.xlsx')
            shutil.copy2(root_file_path, dest_path)
            
            # Deactivate any existing
            BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
            
            wb_obj = BtplWorkbook.objects.create(
                original_name='BTPL_Shipments.xlsx',
                active_sheet='JUN 26',
                uploaded_by=request.user,
                is_active=True
            )
            wb_obj.file.name = 'btpl/BTPL_Shipments.xlsx'
            wb_obj.save(update_fields=['file'])
            
            messages.success(request, "Default BTPL shipment workbook loaded successfully.")
            return redirect('btpl_settings')
            
        elif action == 'change_sheet':
            form = BtplWorkbookUploadForm(request.POST, sheets=sheets)
            if form.is_valid():
                sheet_sel = form.cleaned_data.get('active_sheet')
                if wb_obj:
                    wb_obj.active_sheet = sheet_sel
                    wb_obj.save(update_fields=['active_sheet'])
                else:
                    # Copy root file to media directory and register in database
                    import shutil
                    root_file_path = os.path.join(settings.BASE_DIR, 'BTPL_Shipments.xlsx')
                    target_dir = os.path.join(settings.MEDIA_ROOT, 'btpl')
                    os.makedirs(target_dir, exist_ok=True)
                    dest_path = os.path.join(target_dir, 'BTPL_Shipments.xlsx')
                    shutil.copy2(root_file_path, dest_path)
                    
                    wb_obj = BtplWorkbook.objects.create(
                        original_name='BTPL_Shipments.xlsx',
                        active_sheet=sheet_sel,
                        uploaded_by=request.user,
                        is_active=True
                    )
                    wb_obj.file.name = 'btpl/BTPL_Shipments.xlsx'
                    wb_obj.save(update_fields=['file'])
                    
                messages.success(request, f"Active sheet tab changed to '{sheet_sel}'.")
                return redirect('btpl_sheet')
                
        elif action == 'upload':
            form = BtplWorkbookUploadForm(request.POST, request.FILES, sheets=sheets)
            if form.is_valid():
                upload = request.FILES.get('workbook')
                if upload:
                    # Create new workbook record
                    new_wb = BtplWorkbook.objects.create(
                        file=upload, original_name=upload.name,
                        uploaded_by=request.user, is_active=False
                    )
                    
                    # Read sheets from the uploaded workbook to set default active sheet
                    try:
                        import openpyxl
                        temp_wb = openpyxl.load_workbook(new_wb.file.path, read_only=True)
                        uploaded_sheets = temp_wb.sheetnames
                        default_sheet = uploaded_sheets[0] if uploaded_sheets else 'JUN 26'
                    except Exception:
                        default_sheet = 'JUN 26'
                        
                    new_wb.active_sheet = default_sheet
                    new_wb.save(update_fields=['active_sheet'])
                    
                    # Deactivate existing active workbook
                    BtplWorkbook.objects.filter(is_active=True).update(is_active=False)
                    new_wb.is_active = True
                    new_wb.save(update_fields=['is_active'])
                    
                    messages.success(request, f"Uploaded workbook '{upload.name}' is now the active BTPL shipment workbook.")
                    return redirect('btpl_settings')
                else:
                    messages.error(request, "Please choose a valid .xlsx file to upload.")
    else:
        form = BtplWorkbookUploadForm(initial={'active_sheet': current_sheet_name}, sheets=sheets)
        
    context = {
        'active': 'btpl',
        'form': form,
        'wb': wb_obj,
        'current_sheet': current_sheet_name,
        'sheets': sheets,
    }
    return render(request, 'core/portal/btpl_settings.html', context)


