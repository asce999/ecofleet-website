from django.shortcuts import render, redirect
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from core.models import Pincode, ToolRun, ToolRunFile
from core.decorators import staff_required, director_required
from core.views.ftl import get_active_ftl_workbook
from django.core.cache import cache
from core import ftl as ftl_logic
import datetime
import os


@staff_required
def dashboard(request):
    User = get_user_model()
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ── Pincode coverage ──
    total_pincodes = Pincode.objects.count()
    oda = Pincode.objects.filter(location_type__iexact='ODA').count()
    non_oda = total_pincodes - oda

    # ── Operational counts (light up as the tools come online) ──
    cofs_month = ToolRun.objects.filter(
        tool=ToolRun.TOOL_COF, status=ToolRun.STATUS_SUCCESS,
        created_at__gte=month_start).count()
    reports_month = ToolRun.objects.filter(
        tool__in=[ToolRun.TOOL_MORNING, ToolRun.TOOL_PENDENCY, ToolRun.TOOL_PREV_MONTH],
        status=ToolRun.STATUS_SUCCESS, created_at__gte=month_start).count()
    total_runs = ToolRun.objects.count()
    team_members = User.objects.filter(is_staff=True, is_active=True).count()

    # ── Top states by coverage ──
    top_states = list(
        Pincode.objects.values('state')
        .annotate(c=Count('id')).order_by('-c')[:8])
    state_labels = [s['state'] or 'Unknown' for s in top_states]
    state_data = [s['c'] for s in top_states]

    # ── Tool activity, last 14 days ──
    since = (now - datetime.timedelta(days=13)).date()
    by_day = {
        r['d']: r['c'] for r in
        ToolRun.objects.filter(created_at__date__gte=since)
        .annotate(d=TruncDate('created_at')).values('d')
        .annotate(c=Count('id'))
    }
    activity_labels, activity_data = [], []
    for i in range(13, -1, -1):
        day = (now - datetime.timedelta(days=i)).date()
        activity_labels.append(day.strftime('%d %b'))
        activity_data.append(by_day.get(day, 0))

    # ── Tool usage breakdown ──
    usage = {r['tool']: r['c'] for r in
            ToolRun.objects.values('tool').annotate(c=Count('id'))}
    tool_counts = [usage.get(ToolRun.TOOL_COF, 0),
                    usage.get(ToolRun.TOOL_MORNING, 0),
                    usage.get(ToolRun.TOOL_PENDENCY, 0),
                    usage.get(ToolRun.TOOL_PREV_MONTH, 0)]

    chart_data = {
        'coverage': {'oda': oda, 'non_oda': non_oda},
        'states': {'labels': state_labels, 'data': state_data},
        'activity': {'labels': activity_labels, 'data': activity_data},
        'tools': {'labels': ['COF', 'Morning Report', 'Pendency', 'Previous Month Update'], 'data': tool_counts},
    }

    # ── FTL metrics ──
    ftl_wb_obj, ftl_file_path, ftl_sheet_name = get_active_ftl_workbook()
    ftl_total = 0
    ftl_delivered = 0
    ftl_in_transit = 0
    ftl_vendors = 0
    
    # --- CACHE IMPLEMENTATION ---
    # Cache Invalidation Strategy:
    # 1. We include `ftl_wb_obj.id` so the cache invalidates instantly if a new active workbook is set.
    # 2. We include `os.path.getmtime()` to instantly invalidate if the existing file is edited externally.
    # 3. Timeout is set to 900s (15 mins) as a passive fallback.
    cache_key = None
    if ftl_wb_obj and ftl_file_path and os.path.exists(ftl_file_path):
        mtime = os.path.getmtime(ftl_file_path)
        cache_key = f"ftl_metrics_active_{ftl_wb_obj.id}_{mtime}"
        
    cached_metrics = cache.get(cache_key) if cache_key else None
    
    if cached_metrics:
        ftl_total, ftl_delivered, ftl_in_transit, ftl_vendors = cached_metrics
    elif ftl_file_path and os.path.exists(ftl_file_path):
        try:
            import openpyxl
            wb_ftl = openpyxl.load_workbook(ftl_file_path, read_only=True)
            if ftl_sheet_name in wb_ftl.sheetnames:
                sheet_ftl = wb_ftl[ftl_sheet_name]
                mapping_ftl = ftl_logic.get_column_mapping(sheet_ftl)
                vendors_set = set()
                del_col = mapping_ftl.get('delivery_date')
                etd_col = mapping_ftl.get('etd')
                vendor_col = mapping_ftl.get('vendor')
                
                for r in range(2, sheet_ftl.max_row + 1):
                    # Check if row is actually empty
                    has_data = False
                    for key in ['booking_date', 'lr_number', 'consignee', 'vehicle_number']:
                        col = mapping_ftl.get(key)
                        if col:
                            val = sheet_ftl.cell(row=r, column=col).value
                            if val is not None and str(val).strip() != "":
                                has_data = True
                                break
                    if not has_data:
                        continue
                        
                    ftl_total += 1
                    del_val = sheet_ftl.cell(row=r, column=del_col).value if del_col else None
                    etd_val = sheet_ftl.cell(row=r, column=etd_col).value if etd_col else None
                    
                    status = ftl_logic.derive_status(etd_val, del_val)
                    if status == 'Delivered':
                        ftl_delivered += 1
                    elif status == 'In Transit':
                        ftl_in_transit += 1
                        
                    v_val = sheet_ftl.cell(row=r, column=vendor_col).value if vendor_col else None
                    if v_val is not None and str(v_val).strip() != "":
                        vendors_set.add(str(v_val).strip())
                ftl_vendors = len(vendors_set)
                
                if cache_key:
                    cache.set(cache_key, (ftl_total, ftl_delivered, ftl_in_transit, ftl_vendors), timeout=900)
        except Exception:
            pass

    ctx = {
        'active': 'dashboard',
        'today': datetime.date.today().strftime('%A, %d %B %Y'),
        'total_pincodes': total_pincodes,
        'oda': oda,
        'non_oda': non_oda,
        'states_served': Pincode.objects.values('state').distinct().count(),
        'cofs_month': cofs_month,
        'reports_month': reports_month,
        'total_runs': total_runs,
        'team_members': team_members,
        'chart_data': chart_data,
        'recent_runs': ToolRun.objects.select_related('user')[:8],
        'ftl_total': ftl_total,
        'ftl_delivered': ftl_delivered,
        'ftl_in_transit': ftl_in_transit,
        'ftl_vendors': ftl_vendors,
    }
    return render(request, 'core/portal/dashboard.html', ctx)


@staff_required
@director_required
def portal_users(request):
    from django.contrib.auth.models import User
    from core.models import UserProfile

    users = User.objects.filter(is_staff=True).order_by('username')
    for u in users:
        UserProfile.objects.get_or_create(user=u)

    if request.method == 'POST':
        for u in users:
            if u == request.user:
                continue

            profile = u.profile
            role = request.POST.get(f"role_{u.id}", "Employee")
            profile.role = role

            profile.can_use_cof = f"cof_{u.id}" in request.POST
            profile.can_use_morning = f"morning_{u.id}" in request.POST
            profile.can_use_pendency = f"pendency_{u.id}" in request.POST
            profile.can_use_prev_month = f"prev_month_{u.id}" in request.POST
            profile.can_use_btpl = f"btpl_{u.id}" in request.POST
            profile.can_use_attendance = f"attendance_{u.id}" in request.POST
            profile.can_use_ftl = f"ftl_{u.id}" in request.POST
            profile.save()

        messages.success(request, "User roles and permissions updated successfully.")
        return redirect('portal_users')

    profiles = [u.profile for u in users]
    return render(request, 'core/portal/users.html', {
        'active': 'users',
        'profiles': profiles,
    })


