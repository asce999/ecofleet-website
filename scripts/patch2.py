import os

with open('core/views/attendance.py', 'a') as f:
    f.write('''
@staff_required
@tool_permission_required('attendance')
@never_cache
def salary_calculator(request):
    from core.models import EmployeeSalary
    from decimal import Decimal

    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    
    if not file_path or not os.path.exists(file_path):
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'no_sheet': True,
        })

    manual_year = request.POST.get('manual_year') or request.GET.get('manual_year')
    manual_month = request.POST.get('manual_month') or request.GET.get('manual_month')
    
    data = attendance_logic.get_attendance_data(file_path, sheet_name, manual_year, manual_month)
    if not data:
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'no_sheet': True,
        })
        
    if data.get('requires_manual_date'):
        return render(request, 'core/portal/attendance_salary.html', {
            'active': 'attendance',
            'requires_manual_date': True,
            'sheet_name': sheet_name,
            'wb': wb_obj,
            'sheet_names': data.get('sheet_names', [])
        })
        
    if request.method == 'POST':
        # Save salaries
        for key, val in request.POST.items():
            if key.startswith('salary_'):
                emp_name = key[7:]
                try:
                    if val.strip() == '':
                        salary_val = None
                    else:
                        salary_val = Decimal(val.strip())
                        if salary_val < 0:
                            messages.error(request, f"Salary for {emp_name} cannot be negative.")
                            continue
                            
                    EmployeeSalary.objects.update_or_create(
                        employee_name=emp_name,
                        defaults={'monthly_salary': salary_val}
                    )
                except Exception as e:
                    messages.error(request, f"Invalid salary for {emp_name}: {val}")
                    
        messages.success(request, "Salaries saved successfully.")
        manual_year = request.POST.get('manual_year')
        manual_month = request.POST.get('manual_month')
        redirect_url = reverse('salary_calculator')
        if manual_year and manual_month:
            redirect_url += f"?manual_year={manual_year}&manual_month={manual_month}"
        return redirect(redirect_url)

    # Ensure all employees in the sheet have an EmployeeSalary record (even if None)
    employee_names = [emp['name'] for emp in data['employees']]
    existing_salaries = EmployeeSalary.objects.filter(employee_name__in=employee_names).values_list('employee_name', flat=True)
    for name in employee_names:
        if name not in existing_salaries:
            EmployeeSalary.objects.get_or_create(employee_name=name)

    salary_data = attendance_logic.calculate_salary_data(data)
    
    context = {
        'active': 'attendance',
        'sheet_name': sheet_name,
        'wb': wb_obj,
        'salary_data': salary_data,
        'sheet_names': data['sheet_names'],
        'manual_year': data.get('year'),
        'manual_month': data.get('month'),
    }
    return render(request, 'core/portal/attendance_salary.html', context)


@staff_required
@tool_permission_required('attendance')
def salary_calculator_export(request):
    import io
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb_obj, file_path, sheet_name = attendance_logic.get_active_attendance_workbook()
    if not file_path or not os.path.exists(file_path):
        messages.error(request, "Attendance workbook file not found.")
        return redirect('salary_calculator')

    manual_year = request.GET.get('manual_year')
    manual_month = request.GET.get('manual_month')
    
    data = attendance_logic.get_attendance_data(file_path, sheet_name, manual_year, manual_month)
    if not data or data.get('requires_manual_date'):
        messages.error(request, "Cannot export: Invalid sheet date or data.")
        return redirect('salary_calculator')

    salary_data = attendance_logic.calculate_salary_data(data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary_{sheet_name}"

    headers = ["Employee Name", "Present Days", "Absent Days", "Monthly Salary (Rs)", "Per Day Salary (Rs)", "Salary Deduction (Rs)", "Salary Payable (Rs)"]
    
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border_side = Side(border_style='thin', color='E5E7EB')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center')

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    for i, emp in enumerate(salary_data['employees'], start=2):
        ws.cell(row=i, column=1, value=emp['name']).border = thin_border
        
        c2 = ws.cell(row=i, column=2, value=emp['present_days'])
        c2.border = thin_border
        c2.alignment = center_align
        
        c3 = ws.cell(row=i, column=3, value=emp['absent_days'])
        c3.border = thin_border
        c3.alignment = center_align
        
        c4 = ws.cell(row=i, column=4, value=float(emp['monthly_salary']) if emp['monthly_salary'] is not None else 0.0)
        c4.border = thin_border
        c4.number_format = '#,##0.00'
        
        c5 = ws.cell(row=i, column=5, value=float(emp['per_day_salary']))
        c5.border = thin_border
        c5.number_format = '#,##0.00'
        
        c6 = ws.cell(row=i, column=6, value=float(emp['salary_deduction']))
        c6.border = thin_border
        c6.number_format = '#,##0.00'
        
        c7 = ws.cell(row=i, column=7, value=float(emp['salary_payable']))
        c7.border = thin_border
        c7.font = Font(bold=True)
        c7.number_format = '#,##0.00'

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Salary_{sheet_name}.xlsx"'
    return response
''')
